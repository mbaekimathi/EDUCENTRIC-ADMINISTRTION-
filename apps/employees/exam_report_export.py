import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _safe_sheet_title(value, used):
    title = re.sub(r"[\\/*?:\[\]]", " ", (value or "Sheet")[:31]).strip() or "Sheet"
    base = title
    counter = 2
    while title in used:
        suffix = f" {counter}"
        title = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(title)
    return title


def _safe_filename_part(value):
    cleaned = re.sub(r"[^\w\s-]+", "", (value or "").strip(), flags=re.UNICODE)
    slug = re.sub(r"[-\s]+", "-", cleaned).strip("-").casefold()
    return slug or "exam-report"


def _format_mark_cell(cell, mode):
    if mode == "raw":
        percent = cell.get("percent")
        if percent is not None:
            return str(percent)
        raw = cell.get("raw")
        if raw in (None, ""):
            return ""
        out_of = cell.get("out_of")
        try:
            out_of_value = int(out_of)
        except (TypeError, ValueError):
            out_of_value = None
        if out_of_value not in (None, 100):
            return f"{raw}/{out_of_value}"
        return str(raw)
    percent = cell.get("percent")
    if percent is None:
        return ""
    grade = (cell.get("grade") or "").strip()
    return f"{percent} ({grade})" if grade else str(percent)


def _autosize_columns(worksheet, max_width=42):
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 0
        for cell in column_cells:
            if cell.value in (None, ""):
                continue
            width = max(width, min(len(str(cell.value)), max_width))
        worksheet.column_dimensions[letter].width = max(width + 2, 10)


def _write_meta_rows(worksheet, report, sheet_title):
    worksheet.append([sheet_title])
    worksheet.append([report.get("exam_title") or ""])
    worksheet.append([report.get("scope_label") or ""])
    worksheet.append([])
    for row_index in range(1, 4):
        worksheet.cell(row=row_index, column=1).font = Font(bold=True)


def _write_matrix_sheet(worksheet, report, sheet, mode):
    _write_meta_rows(worksheet, report, sheet.get("exam_title") or "Exam marks")
    header = ["Learner", "Class", "Admission no."]
    header.extend(
        (subject.code or subject.name)
        for subject in sheet.get("subjects") or []
    )
    if mode == "graded":
        header.append("Average")
    worksheet.append(header)
    for row in sheet.get("rows") or []:
        student = row.get("student")
        values = [
            student.display_name if student is not None else "",
            row.get("class_label") or "",
            row.get("admission") or "",
        ]
        values.extend(_format_mark_cell(cell, mode) for cell in row.get("cells") or [])
        if mode == "graded":
            mean = row.get("mean_percent")
            grade = (row.get("overall_grade") or "").strip()
            if mean is None:
                values.append("")
            elif grade:
                values.append(f"{mean} ({grade})")
            else:
                values.append(str(mean))
        worksheet.append(values)
    _autosize_columns(worksheet)


def _write_individual_sheet(worksheet, report, cards, mode):
    _write_meta_rows(worksheet, report, report.get("kind_label") or "Individual report")
    exam_columns = []
    if cards:
        exam_columns = cards[0].get("exam_columns") or []
    header = ["Student", "Admission no.", "Class", "Subject code", "Subject name"]
    header.extend(column.get("label") or column.get("title") or "Assessment" for column in exam_columns)
    if mode == "graded" and exam_columns:
        header.append("Subject average")
    worksheet.append(header)

    selected_class = report.get("selected_class")
    class_label = selected_class.display_label if selected_class is not None else ""

    for card in cards:
        student = card.get("student")
        student_name = student.display_name if student is not None else ""
        admission = (student.admission_number or "") if student is not None else ""
        student_class = class_label or ((student.class_group or "") if student is not None else "")
        for row in card.get("rows") or []:
            subject = row.get("subject")
            values = [
                student_name,
                admission,
                student_class,
                (subject.code or "") if subject is not None else "",
                (subject.name or "") if subject is not None else "",
            ]
            values.extend(_format_mark_cell(cell, mode) for cell in row.get("cells") or [])
            if mode == "graded" and exam_columns:
                mean = row.get("mean_percent")
                grade = (row.get("grade") or "").strip()
                if mean is None:
                    values.append("")
                elif grade:
                    values.append(f"{mean} ({grade})")
                else:
                    values.append(str(mean))
            worksheet.append(values)
    _autosize_columns(worksheet)


def build_exam_report_excel(report, *, mode="raw"):
    if mode not in {"raw", "graded"}:
        mode = "raw"
    workbook = Workbook()
    used_titles = set()
    if report.get("is_matrix"):
        workbook.remove(workbook.active)
        for sheet in report.get("matrix_sheets") or []:
            title = _safe_sheet_title(sheet.get("exam_title"), used_titles)
            worksheet = workbook.create_sheet(title=title)
            _write_matrix_sheet(worksheet, report, sheet, mode)
        if not workbook.sheetnames:
            worksheet = workbook.create_sheet(title="Report")
            _write_matrix_sheet(worksheet, report, {}, mode)
    else:
        worksheet = workbook.active
        worksheet.title = _safe_sheet_title("Individual report", used_titles)
        _write_individual_sheet(worksheet, report, report.get("report_cards") or [], mode)

    buffer = BytesIO()
    workbook.save(buffer)
    issued_on = report.get("issued_on")
    date_part = issued_on.isoformat() if issued_on is not None else "export"
    filename = (
        f"{_safe_filename_part(report.get('exam_title'))}-"
        f"{_safe_filename_part(report.get('scope_label'))}-"
        f"{mode}-{date_part}.xlsx"
    )
    return buffer.getvalue(), filename
