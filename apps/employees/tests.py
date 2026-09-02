from datetime import date, time

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from apps.curriculum.models import (
    AcademicClass,
    AcademicLevel,
    AcademicTerm,
    AcademicYear,
    ClassSubjectAllocation,
    ClassSubjectLessonPlan,
    CombinedExamSubject,
    CombinedExamSubjectComponent,
    ELearningLearningMaterial,
    ELearningSubjectAllocation,
    ELearningSubjectLessonPlan,
    ExamMark,
    ExamScheduleProfile,
    ExamSubjectSetting,
    ExamSupervisorAllocation,
    GeneratedExamSitting,
    GeneratedExamTimetable,
    GeneratedELearningLesson,
    GeneratedELearningTimetable,
    GeneratedLearningLesson,
    GeneratedLearningTimetable,
    GradeBand,
    LearningArea,
    LearningScheduleProfile,
)

from .models import Employee, EmployeeRole, SchoolProfile


class EmployeeAuthenticationTests(TestCase):
    def setUp(self):
        self.employee = Employee.objects.create_user(
            employee_code="123456",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="AMINA",
            last_name="OTIENO",
            email="amina@example.com",
            phone_number="+254700000000",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.pending_employee = Employee.objects.create_user(
            employee_code="654321",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="JOHN",
            last_name="DOE",
            email="john@example.com",
            phone_number="+254711111111",
        )

    def test_login_page_loads(self):
        response = self.client.get(reverse("employees:login"))
        self.assertContains(response, "Welcome back")

    def test_employee_can_sign_in_with_code_and_password(self):
        response = self.client.post(
            reverse("employees:login"),
            {"employee_code": "123456", "password": "ReliablePass456"},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "teacher"}),
        )

    def test_multi_role_employee_must_select_workspace_role(self):
        self.employee.set_roles(
            [Employee.Role.TEACHER, Employee.Role.ACCOUNTANT],
            primary=Employee.Role.TEACHER,
        )
        response = self.client.post(
            reverse("employees:login"),
            {"employee_code": "123456", "password": "ReliablePass456"},
        )
        self.assertRedirects(response, reverse("employees:select_login_role"))
        response = self.client.get(reverse("employees:select_login_role"))
        self.assertContains(response, "Select a role")
        self.assertContains(response, "Teacher")
        self.assertContains(response, "Accountant")
        response = self.client.post(
            reverse("employees:select_login_role"),
            {"role": Employee.Role.ACCOUNTANT},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "accountant"}),
        )
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "accountant"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Switch role")


    def test_registration_creates_employee(self):
        response = self.client.post(
            reverse("employees:register"),
            {
                "title": "MR",
                "first_name": "Grace",
                "last_name": "Doe",
                "email": "grace@example.com",
                "country_code": "KE",
                "phone_national": "0722222222",
                "employee_code": "987654",
                "password1": "ReliablePass456",
                "password2": "ReliablePass456",
            },
        )
        self.assertRedirects(response, reverse("employees:login"))
        employee = Employee.objects.get(employee_code="987654")
        self.assertEqual(employee.first_name, "GRACE")
        self.assertEqual(employee.last_name, "DOE")
        self.assertEqual(employee.email, "grace@example.com")
        self.assertEqual(employee.phone_number, "+254722222222")
        self.assertTrue(employee.check_password("ReliablePass456"))
        self.assertEqual(employee.role, Employee.Role.EMPLOYEE)
        self.assertEqual(
            employee.approval_status,
            Employee.ApprovalStatus.PENDING_APPROVAL,
        )
        self.assertFalse(employee.is_active)
        self.assertEqual(employee.employment_number, 3)

    def test_registration_page_hides_role_and_shows_phone_country(self):
        response = self.client.get(reverse("employees:register"))
        self.assertContains(response, "Phone")
        self.assertContains(response, "country_code")
        self.assertNotContains(response, 'name="role"')
        self.assertNotContains(response, 'name="employment_number"')
        self.assertContains(response, "password-toggle")
        self.assertNotContains(response, "Identity")
        self.assertNotContains(response, "profile_image")

    def test_pending_employee_cannot_sign_in(self):
        response = self.client.post(
            reverse("employees:login"),
            {"employee_code": "654321", "password": "ReliablePass456"},
        )
        self.assertContains(response, "pending administrator approval")

    def test_approving_employee_allows_login(self):
        self.pending_employee.approval_status = Employee.ApprovalStatus.APPROVED
        self.pending_employee.save()
        self.pending_employee.refresh_from_db()
        self.assertTrue(self.pending_employee.is_active)
        response = self.client.post(
            reverse("employees:login"),
            {"employee_code": "654321", "password": "ReliablePass456"},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "employee"}),
        )

    def test_registration_accepts_letters_and_digits_password(self):
        response = self.client.post(
            reverse("employees:register"),
            {
                "title": "MS",
                "first_name": "Nora",
                "last_name": "Kimani",
                "email": "nora@example.com",
                "country_code": "KE",
                "phone_national": "733333333",
                "employee_code": "112233",
                "password1": "school123",
                "password2": "school123",
            },
        )
        self.assertRedirects(response, reverse("employees:login"))
        employee = Employee.objects.get(employee_code="112233")
        self.assertTrue(employee.check_password("school123"))
        self.assertEqual(employee.phone_number, "+254733333333")
        self.assertEqual(employee.role, Employee.Role.EMPLOYEE)


class AcademicCalendarSettingsTests(TestCase):
    def setUp(self):
        self.employee = Employee.objects.create_user(
            employee_code="123456",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="AMINA",
            last_name="OTIENO",
            email="amina@example.com",
            phone_number="+254700000000",
            role=Employee.Role.HEAD_OF_INSTITUTION,
            approval_status=Employee.ApprovalStatus.APPROVED,
        )
        self.client.force_login(self.employee)
        self.url = reverse("employees:academic_calendar_settings")

    def test_system_settings_links_to_academic_calendar(self):
        response = self.client.get(reverse("employees:system_settings"))
        self.assertContains(response, reverse("employees:academic_calendar_settings"))
        self.assertContains(response, "Academic calendar settings")

    def test_academic_year_and_terms_can_be_registered(self):
        response = self.client.post(
            self.url,
            {
                "start_date": "2026-01-05",
                "end_date": "2026-11-20",
                "is_current": "on",
                "status": "ACTIVE",
                "term_id": ["", ""],
                "term_name": ["TERM 1", "TERM 2"],
                "term_start": ["2026-01-05", "2026-05-04"],
                "term_end": ["2026-04-10", "2026-08-07"],
                "term_opening": ["2026-01-12", "2026-05-11"],
                "term_midterm": ["2026-02-20", "2026-06-19"],
                "term_closing": ["2026-04-03", "2026-07-31"],
            },
        )
        self.assertRedirects(response, self.url)
        year = AcademicYear.objects.get(name="2026")
        self.assertTrue(year.is_current)
        self.assertEqual(year.start_date.isoformat(), "2026-01-05")
        terms = list(year.terms.order_by("order"))
        self.assertEqual([term.name for term in terms], ["TERM 1", "TERM 2"])
        self.assertEqual(terms[0].opening_date.isoformat(), "2026-01-12")
        self.assertEqual(terms[0].midterm_date.isoformat(), "2026-02-20")
        self.assertEqual(terms[0].closing_date.isoformat(), "2026-04-03")

    def test_term_dates_must_fall_inside_the_academic_year(self):
        response = self.client.post(
            self.url,
            {
                "start_date": "2026-01-05",
                "end_date": "2026-11-20",
                "status": "ACTIVE",
                "term_id": [""],
                "term_name": ["TERM 1"],
                "term_start": ["2025-12-01"],
                "term_end": ["2026-04-10"],
                "term_opening": ["2026-01-12"],
                "term_midterm": ["2026-02-20"],
                "term_closing": ["2026-04-03"],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Term dates must fall within the academic year.")
        self.assertFalse(AcademicYear.objects.exists())

    def test_current_term_can_be_selected_for_filter_defaults(self):
        year = AcademicYear.objects.create(
            name="2026",
            start_date="2026-01-05",
            end_date="2026-11-20",
            is_current=True,
            status=AcademicYear.Status.ACTIVE,
        )
        term_1 = AcademicTerm.objects.create(
            academic_year=year,
            name="TERM 1",
            start_date="2026-01-05",
            end_date="2026-04-10",
            opening_date="2026-01-12",
            midterm_date="2026-02-20",
            closing_date="2026-04-03",
            order=1,
        )
        term_2 = AcademicTerm.objects.create(
            academic_year=year,
            name="TERM 2",
            start_date="2026-05-04",
            end_date="2026-08-07",
            opening_date="2026-05-11",
            midterm_date="2026-06-19",
            closing_date="2026-07-31",
            order=2,
            is_current=True,
        )

        response = self.client.get(self.url)
        self.assertContains(response, "Current year")
        self.assertContains(response, "Current term")
        self.assertContains(response, reverse("employees:set_current_academic_calendar"))

        response = self.client.post(
            reverse("employees:set_current_academic_calendar"),
            {"year_id": year.id, "term_id": term_1.id},
        )
        self.assertRedirects(response, self.url)
        year.refresh_from_db()
        term_1.refresh_from_db()
        term_2.refresh_from_db()
        self.assertTrue(year.is_current)
        self.assertTrue(term_1.is_current)
        self.assertFalse(term_2.is_current)


class SchoolProfileSettingsTests(TestCase):
    def setUp(self):
        self.employee = Employee.objects.create_user(
            employee_code="123456",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="AMINA",
            last_name="OTIENO",
            email="amina@example.com",
            phone_number="+254700000000",
            role=Employee.Role.HEAD_OF_INSTITUTION,
            approval_status=Employee.ApprovalStatus.APPROVED,
        )
        self.client.force_login(self.employee)

    def test_school_profile_can_be_saved(self):
        response = self.client.post(
            reverse("employees:school_profile_settings"),
            {
                "official_name": "Edu-Centric Academy",
                "display_name": "Edu-Centric",
                "school_type": SchoolProfile.SchoolType.MIXED,
                "ownership": SchoolProfile.Ownership.PRIVATE,
                "moe_code": "MOE-123",
                "nemis_number": "NEMIS-456",
                "knec_centre_number": "KNEC-789",
                "curricula": ["CBC", "IGCSE"],
            },
        )

        self.assertRedirects(response, reverse("employees:school_profile_settings"))
        profile = SchoolProfile.objects.get(pk=1)
        self.assertEqual(profile.official_name, "EDU-CENTRIC ACADEMY")
        self.assertEqual(profile.display_name, "EDU-CENTRIC")
        self.assertEqual(profile.curricula, ["CBC", "IGCSE"])
        self.assertEqual(profile.nemis_number, "NEMIS-456")

    def test_profile_section_pages_are_available_from_sidebar(self):
        response = self.client.get(reverse("employees:school_profile_settings"))
        self.assertContains(response, "school-profile/contact-location")

        pages = (
            "school_profile_contact_location_settings",
            "school_profile_branding_settings",
            "school_profile_leadership_settings",
            "school_profile_academic_setup_settings",
            "school_profile_operations_settings",
            "school_profile_financial_settings",
            "school_profile_compliance_settings",
        )
        for page in pages:
            with self.subTest(page=page):
                response = self.client.get(reverse(f"employees:{page}"))
                self.assertEqual(response.status_code, 200)


class TeacherProfileSettingsTests(TestCase):
    def setUp(self):
        self.teacher = Employee.objects.create_user(
            employee_code="482673",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="ali.teacher@example.com",
            phone_number="+254700000111",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.client.force_login(self.teacher)

    def test_teacher_workspace_shows_profile_settings_not_system_settings(self):
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Profile settings")
        self.assertContains(response, reverse("employees:profile_settings"))
        self.assertNotContains(response, "System settings")
        self.assertNotContains(response, reverse("employees:system_settings"))

    def test_teacher_is_redirected_from_system_settings(self):
        response = self.client.get(reverse("employees:system_settings"))
        self.assertRedirects(response, reverse("employees:profile_settings"))

    def test_teacher_can_update_profile_phone(self):
        response = self.client.post(
            reverse("employees:profile_settings"),
            {
                "form_type": "account",
                "country_code": "KE",
                "phone_national": "711222333",
            },
        )
        self.assertRedirects(response, reverse("employees:profile_settings"))
        self.teacher.refresh_from_db()
        self.assertEqual(self.teacher.phone_number, "+254711222333")

    def test_non_teacher_is_redirected_from_profile_settings(self):
        admin = Employee.objects.create_user(
            employee_code="123456",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="AMINA",
            last_name="OTIENO",
            email="amina@example.com",
            phone_number="+254700000000",
            role=Employee.Role.HEAD_OF_INSTITUTION,
            approval_status=Employee.ApprovalStatus.APPROVED,
        )
        self.client.force_login(admin)
        response = self.client.get(reverse("employees:profile_settings"))
        self.assertRedirects(response, reverse("employees:system_settings"))


class ITSupportWorkspaceTests(TestCase):
    def setUp(self):
        self.employee = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135791",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="preview.teacher@example.com",
            phone_number="+254700000333",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.client.force_login(self.employee)

    def test_it_support_dashboard_shows_module_sidebar_links(self):
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "it_support"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Human resource management")
        self.assertContains(response, "Student management")
        self.assertContains(response, "Curriculum management")
        self.assertContains(response, "Financial management")
        self.assertContains(response, "Stock management")
        self.assertContains(response, "Reports")
        self.assertContains(response, "/workspace/it_support/human-resource-management/")
        self.assertContains(response, "/workspace/it_support/student-management/")
        self.assertContains(response, "/workspace/it_support/curriculum-management/")
        self.assertContains(response, "/workspace/it_support/financial-management/")
        self.assertContains(response, "/workspace/it_support/stock-management/")
        self.assertContains(response, "/workspace/it_support/reports/")

    def test_it_support_dashboard_shows_system_performance_link_and_widget(self):
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "it_support"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "System performance")
        self.assertContains(response, reverse("employees:it_support_system_performance"))
        self.assertContains(response, reverse("employees:it_support_system_performance_metrics"))
        self.assertContains(response, "data-system-performance")
        self.assertContains(response, "Performance trend")
        self.assertContains(response, "Data volumes")
        self.assertContains(response, "sys-perf-kpi__trend")

    def test_it_support_system_performance_page_loads(self):
        response = self.client.get(reverse("employees:it_support_system_performance"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "System performance")
        self.assertContains(response, reverse("employees:it_support_system_performance_metrics"))
        self.assertContains(response, "Live system health")
        self.assertContains(response, "Metric trends")
        self.assertContains(response, "sys-perf-line-chart")
        self.assertContains(response, "data-sys-perf-refresh")

    def test_it_support_system_performance_metrics_returns_json(self):
        response = self.client.get(reverse("employees:it_support_system_performance_metrics"))
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("status", payload)
        self.assertIn("database", payload)
        self.assertIn("cache", payload)
        self.assertIn("storage", payload)
        self.assertIn("counts", payload)
        self.assertIn("operations", payload)
        self.assertIn("tables", payload)
        self.assertIn("latency", payload)
        self.assertIn("media", payload)
        self.assertIn("active_sessions", payload)
        self.assertIn("stress_timeline", payload)
        self.assertIn("kpi_trends", payload)
        self.assertIn("health", payload)
        self.assertIn("trend_cards", payload)
        self.assertIn("chart", payload["latency"])
        self.assertEqual(payload["database"]["status"], "ok")
        self.assertIn("students_total", payload["counts"])
        self.assertIn("students_admitted_this_month", payload["operations"])
        self.assertIn("employees", payload["active_sessions"]["totals"])

    def test_active_sessions_include_logged_in_employee(self):
        from django.contrib.sessions.backends.db import SessionStore
        from django.core.cache import cache

        from apps.employees.live_sessions import LIVE_SESSION_ACTIVITY_KEY
        from apps.employees.system_performance import ACTIVE_SESSIONS_CACHE_KEY
        from apps.employees.workspace import ACTIVE_WORKSPACE_ROLE_SESSION_KEY

        cache.delete(ACTIVE_SESSIONS_CACHE_KEY)

        store = SessionStore()
        store[ACTIVE_WORKSPACE_ROLE_SESSION_KEY] = Employee.Role.IT_SUPPORT
        store["_auth_user_id"] = str(self.employee.pk)
        store[LIVE_SESSION_ACTIVITY_KEY] = timezone.now().isoformat()
        store.save()

        response = self.client.get(reverse("employees:it_support_system_performance_metrics"))
        payload = response.json()
        employee_names = [row["name"] for row in payload["active_sessions"]["employees"]]
        self.assertIn(self.employee.display_name, employee_names)

    def test_active_sessions_exclude_stale_employee_sessions(self):
        from datetime import timedelta

        from django.contrib.sessions.backends.db import SessionStore
        from django.core.cache import cache

        from apps.employees.live_sessions import LIVE_SESSION_ACTIVITY_KEY
        from apps.employees.system_performance import ACTIVE_SESSIONS_CACHE_KEY
        from apps.employees.workspace import ACTIVE_WORKSPACE_ROLE_SESSION_KEY

        cache.delete(ACTIVE_SESSIONS_CACHE_KEY)

        store = SessionStore()
        store[ACTIVE_WORKSPACE_ROLE_SESSION_KEY] = Employee.Role.IT_SUPPORT
        store["_auth_user_id"] = str(self.employee.pk)
        store[LIVE_SESSION_ACTIVITY_KEY] = (timezone.now() - timedelta(hours=2)).isoformat()
        store.save()

        response = self.client.get(reverse("employees:it_support_system_performance_metrics"))
        payload = response.json()
        employee_names = [row["name"] for row in payload["active_sessions"]["employees"]]
        self.assertNotIn(self.employee.display_name, employee_names)

    def test_stress_event_recorded_when_system_is_degraded(self):
        from django.contrib.sessions.backends.db import SessionStore

        from apps.employees.live_sessions import LIVE_SESSION_ACTIVITY_KEY
        from apps.employees.system_performance import ACTIVE_SESSIONS_CACHE_KEY, STRESS_EVENTS_KEY, get_system_performance_snapshot
        from apps.employees.workspace import ACTIVE_WORKSPACE_ROLE_SESSION_KEY
        from django.core.cache import cache

        cache.delete(STRESS_EVENTS_KEY)
        cache.delete(ACTIVE_SESSIONS_CACHE_KEY)
        store = SessionStore()
        store[ACTIVE_WORKSPACE_ROLE_SESSION_KEY] = Employee.Role.IT_SUPPORT
        store["_auth_user_id"] = str(self.employee.pk)
        store[LIVE_SESSION_ACTIVITY_KEY] = timezone.now().isoformat()
        store.save()

        snapshot = get_system_performance_snapshot()
        self.assertIn("stress_timeline", snapshot)
        self.assertTrue(snapshot["stress_timeline"]["events"])

        event = snapshot["stress_timeline"]["events"][0]
        self.assertIn("summary", event)
        self.assertIn("reasons", event)
        self.assertIn("sessions", event)
        self.assertIn(self.employee.display_name, [row["name"] for row in event["sessions"]["employees"]])

    def test_it_support_dashboard_renders_live_snapshot_values(self):
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "it_support"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "sys-perf-initial")
        self.assertNotContains(response, "Checking…")
        self.assertContains(response, "Users in session")
        self.assertContains(response, "ms")

    def test_system_performance_snapshot_includes_kpi_trends(self):
        from apps.employees.system_performance import get_system_performance_snapshot

        snapshot = get_system_performance_snapshot()
        self.assertIn("kpi_trends", snapshot)
        self.assertIn("db", snapshot["kpi_trends"])
        self.assertIn("cache", snapshot["kpi_trends"])
        self.assertIn("disk", snapshot["kpi_trends"])
        self.assertIn("sessions", snapshot["kpi_trends"])
        self.assertIn("score", snapshot["health"])
        self.assertEqual(len(snapshot["trend_cards"]), 4)

    def test_teacher_cannot_access_system_performance_metrics(self):
        response = self.client.get(reverse("employees:it_support_system_performance_metrics"))
        self.assertEqual(response.status_code, 403)

    def test_it_support_module_pages_load(self):
        modules = (
            "human-resource-management",
            "student-management",
            "curriculum-management",
            "financial-management",
            "stock-management",
            "reports",
        )
        for slug in modules:
            with self.subTest(module=slug):
                response = self.client.get(
                    reverse("employees:it_support_module", kwargs={"module": slug})
                )
                self.assertEqual(response.status_code, 200)
                if slug == "curriculum-management":
                    self.assertContains(response, "Learning management")
                    self.assertContains(response, "E-learning management")
                    self.assertContains(response, "Assessment management")
                    self.assertContains(
                        response,
                        "/workspace/it_support/curriculum-management/learning-management/",
                    )
                    self.assertContains(
                        response,
                        "/workspace/it_support/curriculum-management/e-learning-management/",
                    )
                    self.assertContains(
                        response,
                        "/workspace/it_support/curriculum-management/exam-management/",
                    )
                elif slug == "reports":
                    self.assertContains(response, "Curriculum reports")
                    self.assertContains(response, "Financial reports")
                    self.assertContains(response, "Store reports")
                    self.assertContains(
                        response,
                        "/workspace/it_support/reports/curriculum-reports/",
                    )
                    self.assertContains(
                        response,
                        "/workspace/it_support/reports/financial-reports/",
                    )
                    self.assertContains(
                        response,
                        "/workspace/it_support/reports/store-reports/",
                    )
                else:
                    self.assertNotContains(response, "workspace-nav-label")
                    self.assertNotContains(response, "Learning management")
                    self.assertNotContains(response, "Curriculum reports")
                    if slug != "stock-management":
                        self.assertNotContains(response, "/workspace/it_support/stock-management/")
                    self.assertNotContains(response, "/workspace/it_support/reports/")

    def test_report_section_pages_load(self):
        sections = (
            "curriculum-reports",
            "financial-reports",
            "store-reports",
        )
        for slug in sections:
            with self.subTest(section=slug):
                response = self.client.get(
                    reverse(
                        "employees:it_support_report_section",
                        kwargs={"section": slug},
                    )
                )
                self.assertEqual(response.status_code, 200)
                self.assertNotContains(response, "workspace-nav-label")
                if slug == "curriculum-reports":
                    self.assertContains(response, "Learning reports")
                    self.assertContains(response, "Assessment reports")
                    self.assertContains(
                        response,
                        "/workspace/it_support/reports/curriculum-reports/learning-reports/",
                    )
                    self.assertContains(
                        response,
                        "/workspace/it_support/reports/curriculum-reports/exam-reports/",
                    )
                else:
                    self.assertNotContains(
                        response,
                        "/workspace/it_support/reports/curriculum-reports/",
                    )
                    if slug != "financial-reports":
                        self.assertNotContains(
                            response,
                            "/workspace/it_support/reports/financial-reports/",
                        )
                    if slug != "store-reports":
                        self.assertNotContains(
                            response,
                            "/workspace/it_support/reports/store-reports/",
                        )

    def test_curriculum_report_pages_load(self):
        pages = ("learning-reports", "exam-reports")
        for slug in pages:
            with self.subTest(page=slug):
                response = self.client.get(
                    reverse(
                        "employees:it_support_curriculum_report_page",
                        kwargs={"page": slug},
                    )
                )
                self.assertEqual(response.status_code, 200)
                self.assertNotContains(response, "workspace-nav-label")
                if slug == "learning-reports":
                    self.assertContains(response, "Learning reports")
                    self.assertNotContains(
                        response,
                        "/workspace/it_support/reports/curriculum-reports/exam-reports/",
                    )
                if slug == "exam-reports":
                    self.assertContains(response, "Assessment reports")
                    self.assertContains(response, "Generate assessment report")
                    self.assertContains(response, "Academic year")
                    self.assertContains(response, "All assessments")
                    self.assertContains(response, "Academic level report")
                    self.assertContains(response, "Individual report")
                    self.assertContains(response, "Select academic year")
                    self.assertContains(response, "Whole grade")
                    self.assertContains(response, "Per class")
                    self.assertContains(response, "Report scope")
                    self.assertContains(
                        response,
                        "List every student from all classes in this grade together",
                    )
                    self.assertContains(
                        response,
                        "Choose one class, then generate for that class only",
                    )
                    self.assertContains(response, '"id": "all"')
                    self.assertContains(response, 'nav.type === "reload"')
                    self.assertNotContains(response, "View all exams")
                    self.assertNotContains(
                        response,
                        "/workspace/it_support/reports/curriculum-reports/learning-reports/",
                    )

    def test_exam_report_requires_exam_selection(self):
        response = self.client.get(
            reverse(
                "employees:it_support_curriculum_report_page",
                kwargs={"page": "exam-reports"},
            ),
            {"generate": "1", "year_id": "1", "report_kind": "academic_level"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select an assessment to generate the report.")

    def test_exam_report_raw_excel_export(self):
        from datetime import date

        from apps.admissions.models import ParentGuardian, Student

        level = AcademicLevel.objects.create(name="Grade 1", code="G1", order=1)
        academic_class = AcademicClass.objects.create(
            academic_level=level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        subject.academic_levels.add(level)
        year = AcademicYear.objects.create(
            name="2026",
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_current=True,
        )
        term = AcademicTerm.objects.create(
            academic_year=year,
            name="TERM 1",
            start_date=date(2026, 1, 5),
            end_date=date(2026, 4, 1),
            opening_date=date(2026, 1, 5),
            midterm_date=date(2026, 2, 15),
            closing_date=date(2026, 4, 1),
            order=1,
        )
        exam = GeneratedExamTimetable.objects.create(
            academic_year=year,
            academic_term=term,
            start_date=date(2026, 3, 10),
            end_date=date(2026, 3, 12),
        )
        exam.academic_levels.add(level)
        parent = ParentGuardian.objects.create(
            full_name="PAT PARENT",
            relationship_to_student="MOTHER",
            phone_number="+254700009999",
            email="pat.parent@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="9001",
            class_group="G1E",
            assessment_number="A9001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=exam,
            student=student,
            learning_area=subject,
            marks=42,
            out_of_marks=50,
        )
        GradeBand.objects.create(
            academic_level=level,
            code="A",
            meaning="Excellent",
            start_percent=80,
            end_percent=100,
            points=12,
            mark_level="Exceeding",
        )

        report_params = {
            "generate": "1",
            "year_id": str(year.id),
            "exam_id": str(exam.id),
            "report_kind": "academic_level",
            "level_id": str(level.id),
            "level_scope": "individual_class",
            "class_id": str(academic_class.id),
        }
        report_page = self.client.get(
            reverse(
                "employees:it_support_curriculum_report_page",
                kwargs={"page": "exam-reports"},
            ),
            report_params,
        )
        self.assertEqual(report_page.status_code, 200)
        self.assertContains(report_page, "Export Excel (raw marks)")
        self.assertContains(report_page, "Export Excel (graded)")

        response = self.client.get(
            reverse("employees:it_support_exam_report_export"),
            {**report_params, "export_mode": "raw"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

        from io import BytesIO

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        flat = [cell for row in rows for cell in row if cell not in (None, "")]
        self.assertIn("ANN EAST", flat)
        self.assertIn(84, flat)
        self.assertNotIn("42/50", flat)
        self.assertNotIn("84/100", flat)
        self.assertNotIn("A", [str(item) for item in flat if str(item) == "A"])

    def test_student_management_lists_all_students(self):
        from apps.admissions.models import ParentGuardian, Student

        parent = ParentGuardian.objects.create(
            full_name="JANE DOE",
            relationship_to_student="MOTHER",
            phone_number="+254700000222",
            email="jane@example.com",
        )
        Student.objects.create(
            first_name="PRECIOUS",
            last_name="KENDI K",
            date_of_birth="2015-06-06",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_4,
            admission_number="3445",
            class_group="4X",
            assessment_number="B006720544",
            sponsorship_category=Student.SponsorshipCategory.BOTH,
            parent_guardian=parent,
        )
        response = self.client.get(
            reverse("employees:it_support_module", kwargs={"module": "student-management"})
        )
        self.assertContains(response, "Class register")
        self.assertContains(response, "PRECIOUS KENDI K")
        self.assertContains(response, "3445")
        self.assertContains(response, "B006720544")
        self.assertContains(response, "4X")
        self.assertContains(response, "JANE DOE")
        self.assertContains(response, "Edit")
        self.assertContains(response, "Suspend")
        self.assertContains(response, "Delete")

    def test_curriculum_section_pages_load_without_curriculum_sidebar(self):
        response = self.client.get(
            reverse(
                "employees:it_support_curriculum_section",
                kwargs={"section": "learning-management"},
            )
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Class management")
        self.assertContains(response, "Timetable management")
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/class-management/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/timetable-management/",
        )

        response = self.client.get(
            reverse(
                "employees:it_support_curriculum_section",
                kwargs={"section": "exam-management"},
            )
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Allocate supervisors")
        self.assertContains(response, "Generate assessment timetable")
        self.assertContains(response, "All assessments")
        self.assertNotContains(response, "Manual allocation")
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/exam-management/allocate-supervisors/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/exam-management/exam-timetable-generation/?generate=1",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/exam-management/exam-records/",
        )
        self.assertNotContains(
            response,
            "/workspace/it_support/curriculum-management/exam-management/exam-timetable-generation/manual-allocation/",
        )
        self.assertNotContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/class-management/",
        )

        response = self.client.get(
            reverse(
                "employees:it_support_curriculum_section",
                kwargs={"section": "e-learning-management"},
            )
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Allocate subjects")
        self.assertContains(response, "Generate e-learning timetable")
        self.assertContains(response, "E-learning attendance")
        self.assertContains(response, "E-learning assessments")
        self.assertContains(response, "Public learning materials")
        self.assertContains(response, "workspace-nav-label")
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/e-learning-management/allocate-subjects/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/e-learning-management/timetable-generation/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/e-learning-management/attendance/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/e-learning-management/assessments/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/e-learning-management/learning-materials/",
        )

        response = self.client.get(
            reverse("employees:it_support_exam_page", kwargs={"tool": "exam-records"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "All assessments")
        self.assertContains(response, "Registered assessments")
        self.assertContains(response, "No assessments registered yet")
        self.assertNotContains(response, "workspace-nav-label")
        self.assertNotContains(
            response,
            "/workspace/it_support/curriculum-management/exam-management/allocate-supervisors/",
        )

    def test_learning_pages_load_without_learning_sidebar(self):
        response = self.client.get(
            reverse("employees:it_support_learning_page", kwargs={"page": "class-management"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Allocate class teachers")
        self.assertNotContains(response, "Student attendance &amp; progress")
        self.assertNotContains(response, "Teacher attendance &amp; progress")
        self.assertNotContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/class-management/student-attendance-progress/",
        )
        self.assertNotContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/class-management/teacher-attendance-progress/",
        )

        response = self.client.get(
            reverse("employees:it_support_learning_page", kwargs={"page": "timetable-management"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Class and subject allocation")
        self.assertContains(response, "Timetable generation")
        self.assertContains(response, "Timetable analytics")
        self.assertContains(response, "Manual allocation")
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/timetable-management/class-and-subject-allocation/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/timetable-management/timetable-generation/?generate=1",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/timetable-management/manual-allocation/",
        )
        self.assertContains(
            response,
            "/workspace/it_support/curriculum-management/learning-management/timetable-management/timetable-analytics/",
        )

    def test_class_pages_load_without_class_sidebar(self):
        for slug in ("student-attendance-progress", "teacher-attendance-progress"):
            with self.subTest(tool=slug):
                response = self.client.get(
                    reverse("employees:it_support_class_page", kwargs={"tool": slug})
                )
                self.assertEqual(response.status_code, 200)
                self.assertNotContains(response, "workspace-nav-label")
                self.assertNotContains(response, "Allocate class teachers")
                if slug != "student-attendance-progress":
                    self.assertNotContains(
                        response,
                        "/workspace/it_support/curriculum-management/learning-management/class-management/student-attendance-progress/",
                    )

    def test_exam_pages_load_without_shared_exam_sidebar(self):
        for slug in ("allocate-supervisors", "exam-timetable-generation", "exam-records"):
            with self.subTest(tool=slug):
                response = self.client.get(
                    reverse("employees:it_support_exam_page", kwargs={"tool": slug})
                )
                self.assertEqual(response.status_code, 200)
                self.assertNotContains(response, "workspace-nav-label")
                if slug != "allocate-supervisors":
                    self.assertNotContains(
                        response,
                        "/workspace/it_support/curriculum-management/exam-management/allocate-supervisors/",
                    )
                if slug != "exam-records":
                    self.assertNotContains(
                        response,
                        "/workspace/it_support/curriculum-management/exam-management/exam-records/",
                    )

    def test_timetable_pages_load_without_timetable_sidebar(self):
        for slug in (
            "class-and-subject-allocation",
            "timetable-generation",
            "timetable-analytics",
            "manual-allocation",
        ):
            with self.subTest(tool=slug):
                response = self.client.get(
                    reverse("employees:it_support_timetable_page", kwargs={"tool": slug})
                )
                self.assertEqual(response.status_code, 200)
                self.assertNotContains(response, "workspace-nav-label")
                if slug != "class-and-subject-allocation":
                    self.assertNotContains(
                        response,
                        "/workspace/it_support/curriculum-management/learning-management/timetable-management/class-and-subject-allocation/",
                    )

    def test_other_role_dashboards_do_not_show_it_support_module_links(self):
        teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.client.force_login(teacher)
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "/workspace/it_support/human-resource-management/")
        self.assertNotContains(response, "Human resource management")

    def test_hr_module_links_to_employee_management(self):
        response = self.client.get(
            reverse("employees:it_support_module", kwargs={"module": "human-resource-management"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Employee management")
        self.assertContains(
            response,
            "/workspace/it_support/human-resource-management/employee-management/",
        )

    def test_it_support_dashboard_includes_role_switch(self):
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "it_support"})
        )
        self.assertContains(response, "View as")
        self.assertContains(response, reverse("employees:switch_workspace_role"))
        self.assertContains(response, reverse("employees:workspace_role_employees"))
        self.assertContains(response, "Teacher")
        self.assertContains(response, "Head of Institution")
        self.assertContains(response, "Choose the role you want to enter")

    def test_it_support_can_list_employees_for_a_role(self):
        response = self.client.get(
            reverse("employees:workspace_role_employees"),
            {"role": Employee.Role.TEACHER},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["role"], Employee.Role.TEACHER)
        self.assertEqual(
            [employee["id"] for employee in payload["employees"]],
            [self.teacher.id],
        )
        self.assertEqual(payload["employees"][0]["name"], self.teacher.display_name)

    def test_role_switch_requires_an_employee_for_another_role(self):
        response = self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.TEACHER},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "it_support"}),
        )

    def test_it_support_can_view_teacher_pages_as_they_are(self):
        response = self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.TEACHER, "employee_id": self.teacher.id},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "teacher"}),
        )
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Assessment records")
        self.assertContains(response, reverse("employees:teacher_exam_records"))
        self.assertContains(response, "Teacher workspace")
        self.assertContains(response, "View-only session as")
        self.assertContains(response, self.teacher.display_name)
        self.assertNotContains(response, "Human resource management")
        self.employee.refresh_from_db()
        self.assertEqual(self.employee.role, Employee.Role.IT_SUPPORT)

    def test_previewing_another_role_blocks_it_support_modules(self):
        self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.TEACHER, "employee_id": self.teacher.id},
        )
        response = self.client.get(
            reverse("employees:it_support_module", kwargs={"module": "human-resource-management"})
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "teacher"}),
        )

    def test_switching_back_restores_it_support_workspace(self):
        self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.TEACHER, "employee_id": self.teacher.id},
        )
        response = self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.IT_SUPPORT},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "it_support"}),
        )
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "it_support"})
        )
        self.assertContains(response, "Human resource management")

    def test_it_support_with_own_teacher_role_can_switch_without_preview(self):
        self.employee.set_roles(
            [Employee.Role.IT_SUPPORT, Employee.Role.TEACHER],
            primary=Employee.Role.IT_SUPPORT,
        )
        level = AcademicLevel.objects.create(name="Grade 2", code="G2", order=2)
        academic_class = AcademicClass.objects.create(
            academic_level=level,
            name="Grade 2 East",
            code="G2E",
            order=1,
            class_teacher=self.employee,
        )
        subject = LearningArea.objects.create(name="English", code="ENG")
        subject.academic_levels.add(level)
        ClassSubjectAllocation.objects.create(
            academic_class=academic_class,
            learning_area=subject,
            teacher=self.employee,
        )

        response = self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.TEACHER},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "teacher"}),
        )
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "View-only session as")
        self.assertContains(response, "Switch role")
        self.assertContains(response, "Assessment records")

    def test_it_support_own_teacher_role_allows_teacher_posts(self):
        self.employee.set_roles(
            [Employee.Role.IT_SUPPORT, Employee.Role.TEACHER],
            primary=Employee.Role.IT_SUPPORT,
        )
        self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.TEACHER},
        )
        response = self.client.post(reverse("employees:teacher_exam_records"))
        self.assertEqual(response.status_code, 200)

    def test_previewing_teacher_blocks_teacher_posts(self):
        self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.TEACHER, "employee_id": self.teacher.id},
        )
        response = self.client.post(
            reverse("employees:teacher_exam_records"),
            follow=True,
        )
        self.assertContains(
            response,
            "Open this role as yourself for full access.",
        )

    def test_other_roles_cannot_switch_workspace_role(self):
        teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.client.force_login(teacher)
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertNotContains(response, "View as")
        response = self.client.post(
            reverse("employees:switch_workspace_role"),
            {"role": Employee.Role.IT_SUPPORT},
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "teacher"}),
        )


class ExamManagementDashboardTests(TestCase):
    def setUp(self):
        self.employee = Employee.objects.create_user(
            employee_code="246811",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="exam.dashboard@example.com",
            phone_number="+254700000112",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135792",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="exam.dashboard.teacher@example.com",
            phone_number="+254700000334",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(name="Grade 1", code="G1", order=1)
        self.academic_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        self.year = AcademicYear.objects.create(
            name="2026",
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_current=True,
        )
        self.term = AcademicTerm.objects.create(
            academic_year=self.year,
            name="TERM 1",
            start_date=date(2026, 1, 5),
            end_date=date(2026, 4, 1),
            opening_date=date(2026, 1, 5),
            midterm_date=date(2026, 2, 15),
            closing_date=date(2026, 4, 1),
            order=1,
        )
        self.exam = GeneratedExamTimetable.objects.create(
            academic_year=self.year,
            academic_term=self.term,
            start_date=date(2026, 3, 10),
            end_date=date(2026, 3, 12),
            status=GeneratedExamTimetable.Status.IN_SESSION,
        )
        self.exam.academic_levels.add(self.level)
        self.dashboard_url = reverse(
            "employees:it_support_curriculum_section",
            kwargs={"section": "exam-management"},
        )
        self.client.force_login(self.employee)

    def _get_dashboard(self):
        return self.client.get(self.dashboard_url)

    def test_exam_management_dashboard_shows_today_timetable_when_in_session(self):
        today = timezone.localdate()
        GeneratedExamSitting.objects.create(
            generation=self.exam,
            academic_level=self.level,
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
            weekday="MON",
            exam_date=today,
            period_name="Morning",
            start_time=time(8, 0),
            end_time=time(10, 0),
        )

        response = self._get_dashboard()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Current assessment")
        self.assertContains(response, "Today's timetable")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "MATH")
        self.assertContains(response, "In session")
        self.assertNotContains(response, "Class marks analytics")
        self.assertNotContains(response, "Marking progress")

    def test_exam_management_dashboard_hides_timetable_when_marking(self):
        self.exam.status = GeneratedExamTimetable.Status.MARKING
        self.exam.save(update_fields=["status"])

        response = self._get_dashboard()

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Today's timetable")
        self.assertContains(response, "Marking progress")

    def test_exam_management_dashboard_shows_teacher_marking_progress(self):
        from apps.admissions.models import ParentGuardian, Student

        self.exam.status = GeneratedExamTimetable.Status.MARKING
        self.exam.save(update_fields=["status"])
        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700000555",
            email="pat.dashboard@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="G1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=self.exam,
            student=student,
            learning_area=self.subject,
            marks=40,
        )

        response = self._get_dashboard()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Marking progress")
        self.assertContains(response, "ALI TEACHER")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "MATH")
        self.assertContains(response, "1 / 1 marks entered")
        self.assertContains(response, "100%")
        self.assertContains(response, "sys-perf-sparkline")
        self.assertNotContains(response, "Class marks analytics")

    def test_exam_management_dashboard_shows_class_analytics_when_analysing(self):
        from apps.admissions.models import ParentGuardian, Student

        self.exam.status = GeneratedExamTimetable.Status.ANALYSING
        self.exam.save(update_fields=["status"])
        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700000556",
            email="pat.analyse@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1002",
            class_group="G1E",
            assessment_number="A1002",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=self.exam,
            student=student,
            learning_area=self.subject,
            marks=80,
            out_of_marks=100,
        )

        response = self._get_dashboard()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Marks input analytics")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "100% entered")
        self.assertContains(response, "Mean 80%")

    def test_exam_management_dashboard_shows_published_results(self):
        from apps.admissions.models import ParentGuardian, Student

        self.exam.status = GeneratedExamTimetable.Status.PUBLISHED
        self.exam.save(update_fields=["status"])
        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700000557",
            email="pat.publish@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1003",
            class_group="G1E",
            assessment_number="A1003",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=self.exam,
            student=student,
            learning_area=self.subject,
            marks=75,
            out_of_marks=100,
        )

        response = self._get_dashboard()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Published results")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "Assessment reports")
        self.assertContains(response, "75")


class EmployeeManagementTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.accountant = Employee.objects.create_user(
            employee_code="112233",
            password="ReliablePass456",
            title=Employee.Title.MRS,
            first_name="GRACE",
            last_name="FINANCE",
            email="accounts@example.com",
            phone_number="+254700000333",
            role=Employee.Role.ACCOUNTANT,
            approval_status=Employee.ApprovalStatus.PENDING_APPROVAL,
        )
        self.client.force_login(self.support)

    def test_lists_employees_in_single_directory(self):
        response = self.client.get(reverse("employees:it_support_employee_management"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Teacher")
        self.assertContains(response, "IT Support")
        self.assertContains(response, "Accountant")
        self.assertContains(response, "ALI TEACHER")
        self.assertContains(response, "GRACE FINANCE")
        self.assertContains(response, "135790")
        self.assertContains(response, self.teacher.employment_number)
        self.assertContains(response, "Employment number")
        self.assertContains(response, "Employee code")
        self.assertNotContains(response, "employees in this role")
        self.assertNotContains(response, "Jump to role")

    def test_other_roles_cannot_open_employee_management(self):
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("employees:it_support_employee_management"))
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "teacher"}),
        )

    def test_can_edit_employee_details(self):
        response = self.client.post(
            reverse("employees:update_workspace_employee", kwargs={"employee_id": self.teacher.id}),
            {
                "title": "DR",
                "first_name": "ALI",
                "last_name": "KARIUKI",
                "email": "ali.kariuki@example.com",
                "phone_number": "+254700000999",
                "employee_code": "135790",
                "employment_number": "99",
                "role": Employee.Role.CURRICULUM_COORDINATOR,
                "roles": [Employee.Role.CURRICULUM_COORDINATOR],
                "approval_status": Employee.ApprovalStatus.APPROVED,
            },
        )
        self.assertRedirects(response, reverse("employees:it_support_employee_management"))
        self.teacher.refresh_from_db()
        self.assertEqual(self.teacher.last_name, "KARIUKI")
        self.assertEqual(self.teacher.email, "ali.kariuki@example.com")
        self.assertEqual(self.teacher.role, Employee.Role.CURRICULUM_COORDINATOR)
        self.assertEqual(self.teacher.role_values(), [Employee.Role.CURRICULUM_COORDINATOR])
        self.assertEqual(self.teacher.title, Employee.Title.DR)
        self.assertEqual(self.teacher.employment_number, 99)

    def test_can_assign_multiple_roles(self):
        response = self.client.post(
            reverse("employees:update_workspace_employee", kwargs={"employee_id": self.teacher.id}),
            {
                "title": "MR",
                "first_name": "ALI",
                "last_name": "TEACHER",
                "email": "teacher@example.com",
                "phone_number": "+254700000222",
                "employee_code": "135790",
                "employment_number": self.teacher.employment_number,
                "role": Employee.Role.TEACHER,
                "roles": [Employee.Role.TEACHER, Employee.Role.ACCOUNTANT],
                "approval_status": Employee.ApprovalStatus.APPROVED,
            },
        )
        self.assertRedirects(response, reverse("employees:it_support_employee_management"))
        self.teacher.refresh_from_db()
        self.assertEqual(self.teacher.role, Employee.Role.TEACHER)
        self.assertCountEqual(
            self.teacher.role_values(),
            [Employee.Role.TEACHER, Employee.Role.ACCOUNTANT],
        )
        self.assertEqual(
            EmployeeRole.objects.filter(employee=self.teacher).count(),
            2,
        )
        response = self.client.get(reverse("employees:it_support_employee_management"))
        self.assertContains(response, "Accountant")
        self.assertContains(response, "Teacher")
        # Multi-role employee appears once in the flat directory.
        self.assertContains(response, "ALI TEACHER", count=1)

    def test_partial_update_without_roles_preserves_existing_roles(self):
        self.teacher.set_roles(
            [Employee.Role.TEACHER, Employee.Role.ACCOUNTANT],
            primary=Employee.Role.TEACHER,
        )
        response = self.client.post(
            reverse("employees:update_workspace_employee", kwargs={"employee_id": self.teacher.id}),
            {
                "title": self.teacher.title,
                "first_name": self.teacher.first_name,
                "last_name": "KARIUKI",
                "email": self.teacher.email,
                "phone_number": self.teacher.phone_number,
                "employee_code": self.teacher.employee_code,
                "employment_number": str(self.teacher.employment_number),
                "role": Employee.Role.TEACHER,
                "approval_status": self.teacher.approval_status,
            },
        )
        self.assertRedirects(response, reverse("employees:it_support_employee_management"))
        self.teacher.refresh_from_db()
        self.assertEqual(self.teacher.last_name, "KARIUKI")
        self.assertCountEqual(
            self.teacher.role_values(),
            [Employee.Role.TEACHER, Employee.Role.ACCOUNTANT],
        )

    def test_can_suspend_and_unsuspend_employee(self):
        response = self.client.post(
            reverse(
                "employees:toggle_workspace_employee_status",
                kwargs={"employee_id": self.teacher.id},
            )
        )
        self.assertRedirects(response, reverse("employees:it_support_employee_management"))
        self.teacher.refresh_from_db()
        self.assertTrue(self.teacher.is_suspended)
        self.assertFalse(self.teacher.is_active)

        self.client.logout()
        response = self.client.post(
            reverse("employees:login"),
            {"employee_code": "135790", "password": "ReliablePass456"},
        )
        self.assertContains(response, "suspended")

        self.client.force_login(self.support)
        self.client.post(
            reverse(
                "employees:toggle_workspace_employee_status",
                kwargs={"employee_id": self.teacher.id},
            )
        )
        self.teacher.refresh_from_db()
        self.assertFalse(self.teacher.is_suspended)
        self.assertTrue(self.teacher.is_active)

    def test_cannot_suspend_own_account(self):
        self.client.post(
            reverse(
                "employees:toggle_workspace_employee_status",
                kwargs={"employee_id": self.support.id},
            )
        )
        self.support.refresh_from_db()
        self.assertFalse(self.support.is_suspended)
        self.assertTrue(self.support.is_active)

    def test_can_delete_employee(self):
        teacher_id = self.teacher.id
        response = self.client.post(
            reverse("employees:delete_workspace_employee", kwargs={"employee_id": teacher_id})
        )
        self.assertRedirects(response, reverse("employees:it_support_employee_management"))
        self.assertFalse(Employee.objects.filter(pk=teacher_id).exists())

    def test_cannot_delete_own_account(self):
        self.client.post(
            reverse("employees:delete_workspace_employee", kwargs={"employee_id": self.support.id})
        )
        self.assertTrue(Employee.objects.filter(pk=self.support.id).exists())

    def test_cannot_reuse_another_employees_number(self):
        response = self.client.post(
            reverse("employees:update_workspace_employee", kwargs={"employee_id": self.teacher.id}),
            {
                "title": self.teacher.title,
                "first_name": self.teacher.first_name,
                "last_name": self.teacher.last_name,
                "email": self.teacher.email,
                "phone_number": self.teacher.phone_number,
                "employee_code": self.teacher.employee_code,
                "employment_number": str(self.support.employment_number),
                "role": self.teacher.role,
                "approval_status": self.teacher.approval_status,
            },
        )
        self.assertRedirects(response, reverse("employees:it_support_employee_management"))
        self.teacher.refresh_from_db()
        self.assertEqual(self.teacher.employment_number, 2)

    def test_can_override_another_employees_number_on_confirmation(self):
        support_number = self.support.employment_number
        response = self.client.post(
            reverse("employees:update_workspace_employee", kwargs={"employee_id": self.teacher.id}),
            {
                "title": self.teacher.title,
                "first_name": self.teacher.first_name,
                "last_name": self.teacher.last_name,
                "email": self.teacher.email,
                "phone_number": self.teacher.phone_number,
                "employee_code": self.teacher.employee_code,
                "employment_number": str(support_number),
                "role": self.teacher.role,
                "roles": [self.teacher.role],
                "approval_status": self.teacher.approval_status,
                "override_employment_number": "1",
            },
        )
        self.assertRedirects(response, reverse("employees:it_support_employee_management"))
        self.teacher.refresh_from_db()
        self.support.refresh_from_db()
        self.assertEqual(self.teacher.employment_number, support_number)
        self.assertIsNone(self.support.employment_number)

    def test_deleted_employment_number_is_not_reused(self):
        used_number = self.teacher.employment_number
        self.teacher.delete()
        replacement = Employee.objects.create_user(
            employee_code="246811",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="NEW",
            last_name="STAFF",
            email="new.staff@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
        )
        self.assertNotEqual(replacement.employment_number, used_number)
        self.assertEqual(replacement.employment_number, 4)


class StudentManagementTests(TestCase):
    def setUp(self):
        from apps.admissions.models import ParentGuardian, Student

        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.parent = ParentGuardian.objects.create(
            full_name="JANE DOE",
            relationship_to_student="MOTHER",
            phone_number="+254700000222",
            email="jane@example.com",
        )
        self.student = Student.objects.create(
            first_name="PRECIOUS",
            last_name="KENDI",
            date_of_birth="2015-06-06",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_4,
            admission_number="3445",
            class_group="4X",
            assessment_number="B006720544",
            sponsorship_category=Student.SponsorshipCategory.BOTH,
            sponsor_details="ELIZABETH KARIUKI",
            parent_guardian=self.parent,
            home_address="MERU",
        )
        self.client.force_login(self.support)

    def test_can_edit_student_details(self):
        from apps.admissions.models import Student

        response = self.client.post(
            reverse("employees:update_workspace_student", kwargs={"student_id": self.student.id}),
            {
                "first_name": "Precious",
                "last_name": "Kendi K",
                "date_of_birth": "2015-06-06",
                "gender": Student.Gender.FEMALE,
                "academic_level": Student.AcademicLevel.GRADE_5,
                "admission_number": "3445",
                "class_group": "5X",
                "assessment_number": "B006720544",
                "previous_school": "",
                "sponsorship_category": Student.SponsorshipCategory.BOTH,
                "sponsor_details": "Elizabeth Kariuki",
                "parent_guardian_name": "Elizabeth Kariuki",
                "relationship_to_student": "Mother",
                "parent_phone": "+254721665936",
                "parent_email": "elizabeth.kariuki@gmail.com",
                "home_address": "Thangatha, Meru County",
                "medical_notes": "",
                "special_needs": "",
                "emergency_contact": "+254721665936",
            },
        )
        self.assertRedirects(
            response,
            reverse("employees:it_support_module", kwargs={"module": "student-management"}),
        )
        self.student.refresh_from_db()
        self.parent.refresh_from_db()
        self.assertEqual(self.student.last_name, "KENDI K")
        self.assertEqual(self.student.academic_level, Student.AcademicLevel.GRADE_5)
        self.assertEqual(self.student.class_group, "5X")
        self.assertEqual(self.student.home_address, "THANGATHA, MERU COUNTY")
        self.assertEqual(self.parent.full_name, "ELIZABETH KARIUKI")
        self.assertEqual(self.parent.phone_number, "+254721665936")

    def test_can_suspend_and_unsuspend_student(self):
        response = self.client.post(
            reverse(
                "employees:toggle_workspace_student_status",
                kwargs={"student_id": self.student.id},
            )
        )
        self.assertRedirects(
            response,
            reverse("employees:it_support_module", kwargs={"module": "student-management"}),
        )
        self.student.refresh_from_db()
        self.assertTrue(self.student.is_suspended)
        self.assertFalse(self.student.is_active)

        self.client.post(
            reverse(
                "employees:toggle_workspace_student_status",
                kwargs={"student_id": self.student.id},
            )
        )
        self.student.refresh_from_db()
        self.assertFalse(self.student.is_suspended)
        self.assertTrue(self.student.is_active)

    def test_can_delete_student(self):
        from apps.admissions.models import ParentGuardian, Student

        student_id = self.student.id
        parent_id = self.parent.id
        response = self.client.post(
            reverse("employees:delete_workspace_student", kwargs={"student_id": student_id})
        )
        self.assertRedirects(
            response,
            reverse("employees:it_support_module", kwargs={"module": "student-management"}),
        )
        self.assertFalse(Student.objects.filter(pk=student_id).exists())
        self.assertFalse(ParentGuardian.objects.filter(pk=parent_id).exists())

    def test_can_view_student_profile(self):
        list_response = self.client.get(
            reverse("employees:it_support_module", kwargs={"module": "student-management"})
        )
        profile_url = reverse(
            "employees:workspace_student_profile",
            kwargs={"student_id": self.student.id},
        )
        self.assertEqual(profile_url, f"/workspace/students/{self.student.id}/profile/")
        self.assertContains(list_response, profile_url)

        response = self.client.get(profile_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "PRECIOUS KENDI")
        self.assertContains(response, "3445")
        self.assertContains(response, "B006720544")
        self.assertContains(response, "JANE DOE")
        self.assertContains(response, "MOTHER")
        self.assertContains(response, "Personal details")
        self.assertContains(response, "Health and support")
        self.assertContains(response, "Back to register")
        self.assertContains(response, "Student management")

        legacy = self.client.get(
            reverse(
                "employees:workspace_student_profile_legacy",
                kwargs={"student_id": self.student.id},
            )
        )
        self.assertRedirects(legacy, profile_url)

    def test_any_role_can_view_shared_student_profile(self):
        teacher = Employee.objects.create_user(
            employee_code="246813",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="ali.teacher.profile@example.com",
            phone_number="+254700000813",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.client.force_login(teacher)
        profile_url = reverse(
            "employees:workspace_student_profile",
            kwargs={"student_id": self.student.id},
        )
        response = self.client.get(profile_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "PRECIOUS KENDI")
        self.assertContains(response, "Back to dashboard")
        self.assertContains(response, "Assessment records")
        self.assertNotContains(response, "Back to register")
        self.assertNotContains(response, "Human resource management")

    def test_student_search_suggests_by_name_and_numbers(self):
        search_url = reverse("employees:workspace_student_search")
        dashboard = self.client.get(reverse("employees:role_dashboard", kwargs={"role": "it_support"}))
        self.assertContains(dashboard, "Search for student")
        self.assertContains(dashboard, search_url)

        by_name = self.client.get(search_url, {"q": "PRECIOUS"})
        self.assertEqual(by_name.status_code, 200)
        payload = by_name.json()["students"]
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["id"], self.student.id)
        self.assertEqual(payload[0]["assessment_number"], "B006720544")
        self.assertIn(
            reverse(
                "employees:workspace_student_profile",
                kwargs={"student_id": self.student.id},
            ),
            payload[0]["profile_url"],
        )

        by_admission = self.client.get(search_url, {"q": "3445"})
        self.assertEqual(by_admission.json()["students"][0]["id"], self.student.id)

        by_assessment = self.client.get(search_url, {"q": "B006720544"})
        self.assertEqual(by_assessment.json()["students"][0]["id"], self.student.id)

        empty = self.client.get(search_url, {"q": "ZZZNOMATCH"})
        self.assertEqual(empty.json()["students"], [])


class ClassSubjectAllocationTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.accountant = Employee.objects.create_user(
            employee_code="111222",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="JOY",
            last_name="ACCOUNTS",
            email="accounts@example.com",
            phone_number="+254700000333",
            role=Employee.Role.ACCOUNTANT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.academic_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        self.client.force_login(self.support)

    def test_allocation_page_lists_level_subjects_and_teachers(self):
        response = self.client.get(
            reverse(
                "employees:it_support_timetable_page",
                kwargs={"tool": "class-and-subject-allocation"},
            )
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Grade 1")
        self.assertContains(response, "Mathematics")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "ALI TEACHER")
        self.assertNotContains(response, "135790")
        self.assertNotContains(response, "JOY ACCOUNTS")

    def test_multi_role_teacher_with_non_teacher_primary_appears_in_allocation(self):
        multi_role_teacher = Employee.objects.create_user(
            employee_code="864200",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="NELLY",
            last_name="MWEBIA",
            email="nelly.mwebia@example.com",
            phone_number="+254700000555",
            role=Employee.Role.ACCOUNTANT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        multi_role_teacher.set_roles(
            [Employee.Role.ACCOUNTANT, Employee.Role.TEACHER],
            primary=Employee.Role.ACCOUNTANT,
        )

        response = self.client.get(
            reverse(
                "employees:it_support_timetable_page",
                kwargs={"tool": "class-and-subject-allocation"},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "NELLY MWEBIA")
        self.assertNotContains(response, "JOY ACCOUNTS")

    def test_teacher_can_be_allocated_a_subject_in_a_class(self):
        response = self.client.post(
            reverse(
                "employees:it_support_timetable_page",
                kwargs={"tool": "class-and-subject-allocation"},
            ),
            {
                "level_id": str(self.level.id),
                f"teacher_{self.academic_class.id}_{self.subject.id}": str(self.teacher.id),
            },
        )
        self.assertRedirects(
            response,
            reverse(
                "employees:it_support_timetable_page",
                kwargs={"tool": "class-and-subject-allocation"},
            ),
        )
        allocation = ClassSubjectAllocation.objects.get(
            academic_class=self.academic_class,
            learning_area=self.subject,
        )
        self.assertEqual(allocation.teacher, self.teacher)

    def test_reassigning_teacher_keeps_lessons_and_plans_and_marks_collisions(self):
        other_teacher = Employee.objects.create_user(
            employee_code="975310",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="TEACHER",
            email="beth.teacher@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        allocation = ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            teacher=other_teacher,
        )
        lesson_plan = ClassSubjectLessonPlan.objects.create(
            allocation=allocation,
            strand="Numbers",
        )
        generation = GeneratedLearningTimetable.objects.create(created_by=self.support)
        generation.academic_levels.add(self.level)
        east_lesson = GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        west_lesson = GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=other_class,
            learning_area=self.subject,
            teacher=other_teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )

        response = self.client.post(
            reverse(
                "employees:it_support_timetable_page",
                kwargs={"tool": "class-and-subject-allocation"},
            ),
            {
                "level_id": str(self.level.id),
                f"teacher_{self.academic_class.id}_{self.subject.id}": str(other_teacher.id),
                f"teacher_{other_class.id}_{self.subject.id}": str(other_teacher.id),
            },
        )
        self.assertRedirects(
            response,
            reverse(
                "employees:it_support_timetable_page",
                kwargs={"tool": "class-and-subject-allocation"},
            ),
        )

        allocation.refresh_from_db()
        lesson_plan.refresh_from_db()
        east_lesson.refresh_from_db()
        west_lesson.refresh_from_db()
        self.assertEqual(allocation.teacher, other_teacher)
        self.assertEqual(lesson_plan.strand, "Numbers")
        self.assertEqual(ClassSubjectLessonPlan.objects.filter(pk=lesson_plan.pk).count(), 1)
        self.assertEqual(GeneratedLearningLesson.objects.count(), 2)
        self.assertEqual(east_lesson.teacher, other_teacher)
        self.assertEqual(west_lesson.teacher, other_teacher)

        page = self.client.get(
            reverse(
                "employees:it_support_timetable_page",
                kwargs={"tool": "timetable-generation"},
            )
        )
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, 'class="is-collision"')
        self.assertContains(page, str(other_teacher.employment_number))


class ELearningSubjectAllocationTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.accountant = Employee.objects.create_user(
            employee_code="111222",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="JOY",
            last_name="ACCOUNTS",
            email="accounts@example.com",
            phone_number="+254700000333",
            role=Employee.Role.ACCOUNTANT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        self.client.force_login(self.support)

    def test_allocation_page_lists_level_subjects_and_teachers(self):
        response = self.client.get(
            reverse(
                "employees:it_support_elearning_page",
                kwargs={"page": "allocate-subjects"},
            )
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Grade 1")
        self.assertContains(response, "Mathematics")
        self.assertContains(response, "ALI TEACHER")
        self.assertContains(response, "Allocate subjects")
        self.assertContains(response, "workspace-nav-label")
        self.assertNotContains(response, "JOY ACCOUNTS")

    def test_teacher_can_be_allocated_a_subject_for_elearning(self):
        response = self.client.post(
            reverse(
                "employees:it_support_elearning_page",
                kwargs={"page": "allocate-subjects"},
            ),
            {
                "level_id": str(self.level.id),
                f"teacher_{self.level.id}_{self.subject.id}": str(self.teacher.id),
            },
        )
        self.assertRedirects(
            response,
            reverse(
                "employees:it_support_elearning_page",
                kwargs={"page": "allocate-subjects"},
            ),
        )
        allocation = ELearningSubjectAllocation.objects.get(
            academic_level=self.level,
            learning_area=self.subject,
        )
        self.assertEqual(allocation.teacher, self.teacher)

    def test_reassigning_elearning_teacher_keeps_sessions_and_marks_collisions(self):
        other_teacher = Employee.objects.create_user(
            employee_code="975310",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="TEACHER",
            email="beth.teacher@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        science = LearningArea.objects.create(name="Science", code="SCI")
        science.academic_levels.add(self.level)
        ELearningSubjectAllocation.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ELearningSubjectAllocation.objects.create(
            academic_level=self.level,
            learning_area=science,
            teacher=other_teacher,
        )
        generation = GeneratedELearningTimetable.objects.create(created_by=self.support)
        generation.academic_levels.add(self.level)
        math_lesson = GeneratedELearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        science_lesson = GeneratedELearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            learning_area=science,
            teacher=other_teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )

        response = self.client.post(
            reverse(
                "employees:it_support_elearning_page",
                kwargs={"page": "allocate-subjects"},
            ),
            {
                "level_id": str(self.level.id),
                f"teacher_{self.level.id}_{self.subject.id}": str(other_teacher.id),
                f"teacher_{self.level.id}_{science.id}": str(other_teacher.id),
            },
        )
        self.assertRedirects(
            response,
            reverse(
                "employees:it_support_elearning_page",
                kwargs={"page": "allocate-subjects"},
            ),
        )

        math_lesson.refresh_from_db()
        science_lesson.refresh_from_db()
        self.assertEqual(GeneratedELearningLesson.objects.count(), 2)
        self.assertEqual(math_lesson.teacher, other_teacher)
        self.assertEqual(science_lesson.teacher, other_teacher)

        page = self.client.get(
            reverse(
                "employees:it_support_elearning_page",
                kwargs={"page": "timetable-generation"},
            )
        )
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, 'class="is-collision"')
        self.assertContains(page, str(other_teacher.employment_number))


class ClassTeacherAllocationTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.accountant = Employee.objects.create_user(
            employee_code="111222",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="JOY",
            last_name="ACCOUNTS",
            email="accounts@example.com",
            phone_number="+254700000333",
            role=Employee.Role.ACCOUNTANT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.academic_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        self.client.force_login(self.support)

    def test_class_management_page_lists_classes_and_teachers(self):
        response = self.client.get(
            reverse("employees:it_support_learning_page", kwargs={"page": "class-management"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Allocate class teachers")
        self.assertContains(response, "Grade 1")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "ALI TEACHER")
        self.assertNotContains(response, "135790")
        self.assertNotContains(response, "JOY ACCOUNTS")

    def test_teacher_can_be_allocated_as_class_teacher(self):
        response = self.client.post(
            reverse("employees:it_support_learning_page", kwargs={"page": "class-management"}),
            {
                "level_id": str(self.level.id),
                f"teacher_{self.academic_class.id}": str(self.teacher.id),
            },
        )
        self.assertRedirects(
            response,
            reverse("employees:it_support_learning_page", kwargs={"page": "class-management"}),
        )
        self.academic_class.refresh_from_db()
        self.assertEqual(self.academic_class.class_teacher, self.teacher)

    def test_class_teacher_can_be_unassigned(self):
        self.academic_class.class_teacher = self.teacher
        self.academic_class.save(update_fields=["class_teacher"])
        response = self.client.post(
            reverse("employees:it_support_learning_page", kwargs={"page": "class-management"}),
            {
                "level_id": str(self.level.id),
                f"teacher_{self.academic_class.id}": "",
            },
        )
        self.assertRedirects(
            response,
            reverse("employees:it_support_learning_page", kwargs={"page": "class-management"}),
        )
        self.academic_class.refresh_from_db()
        self.assertIsNone(self.academic_class.class_teacher)


class TimetableAnalyticsTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        level = AcademicLevel.objects.create(name="Grade 1", code="G1", category="PRIMARY", order=1)
        academic_class = AcademicClass.objects.create(
            academic_level=level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        subject.academic_levels.add(level)
        ClassSubjectAllocation.objects.create(
            academic_class=academic_class,
            learning_area=subject,
            teacher=self.teacher,
        )
        self.client.force_login(self.support)

    def test_analytics_page_shows_period_teacher_and_subject_sections(self):
        response = self.client.get(
            reverse("employees:it_support_timetable_page", kwargs={"tool": "timetable-analytics"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Learning period allocations")
        self.assertContains(response, "Employee (teacher) allocations")
        self.assertContains(response, "Subject allocations")
        self.assertContains(response, "ALI TEACHER")
        self.assertContains(response, "Mathematics")
        self.assertContains(response, "Grade 1 East")


class TimetableGenerationTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.academic_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        self.client.force_login(self.support)
        self.url = reverse(
            "employees:it_support_timetable_page",
            kwargs={"tool": "timetable-generation"},
        )

    def test_generation_page_opens_popup_and_blocks_unallocated_levels(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-open-modal="timetable-generation"')
        self.assertContains(response, "Grade 1")
        self.assertContains(response, "is-blocked")

    def test_fully_allocated_level_is_blocked_without_timetable_settings(self):
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        response = self.client.get(self.url)
        self.assertContains(response, "is-blocked")
        self.assertContains(response, "learning timetable settings")

    def test_fully_allocated_level_with_timetable_settings_can_be_generated(self):
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY DAY",
            category="PRIMARY",
            study_days=["MON", "WED"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(8, 40),
        )
        profile.academic_levels.add(self.level)
        response = self.client.get(self.url)
        self.assertContains(response, "is-viable")
        self.assertContains(response, "PRIMARY DAY")

        response = self.client.post(self.url, {"level_id": str(self.level.id)})
        self.assertRedirects(response, self.url)
        generation = GeneratedLearningTimetable.objects.get()
        self.assertIn(self.level, generation.academic_levels.all())
        lessons = list(GeneratedLearningLesson.objects.filter(generation=generation).order_by("weekday"))
        self.assertEqual(len(lessons), 2)
        self.assertEqual({lesson.weekday for lesson in lessons}, {"MON", "WED"})
        self.assertEqual(lessons[0].start_time, time(8, 0))
        self.assertEqual(lessons[0].end_time, time(8, 40))
        page = self.client.get(self.url)
        self.assertContains(page, "Grade 1 East")
        self.assertContains(page, "MATH")
        self.assertContains(page, str(self.teacher.employment_number))
        self.assertNotContains(page, "ALI TEACHER")
        self.assertContains(page, "Monday")
        self.assertContains(page, "Wednesday")

    def test_generation_leaves_blank_when_the_same_teacher_would_collide(self):
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY DAY",
            category="PRIMARY",
            study_days=["MON"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(8, 40),
        )
        profile.academic_levels.add(self.level)

        response = self.client.post(self.url, {"level_id": str(self.level.id)})
        self.assertRedirects(response, self.url)
        lessons = list(GeneratedLearningLesson.objects.all())
        self.assertEqual(len(lessons), 1)
        self.assertEqual(lessons[0].teacher, self.teacher)

    def test_generate_query_opens_generation_popup(self):
        response = self.client.get(f"{self.url}?generate=1")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-modal="timetable-generation"')
        self.assertContains(response, "is-open")

    def test_reset_query_opens_reset_popup(self):
        response = self.client.get(f"{self.url}?reset=1")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-modal="timetable-reset"')
        self.assertContains(response, "is-open")

    def test_reset_removes_generated_lessons_for_selected_levels(self):
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY DAY",
            category="PRIMARY",
            study_days=["MON", "WED"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(8, 40),
        )
        profile.academic_levels.add(self.level)
        self.client.post(self.url, {"level_id": str(self.level.id)})
        self.assertEqual(GeneratedLearningLesson.objects.count(), 2)

        response = self.client.post(
            self.url,
            {"action": "reset", "level_id": str(self.level.id)},
        )
        self.assertRedirects(response, self.url)
        self.assertEqual(GeneratedLearningLesson.objects.count(), 0)
        page = self.client.get(self.url)
        self.assertContains(page, "No timetable generated yet")
        self.assertContains(page, 'name="action" value="generate_class"')
        self.assertRegex(
            page.content.decode(),
            r'data-open-modal="timetable-reset"[^>]*disabled|disabled[^>]*data-open-modal="timetable-reset"',
        )

    def test_generate_class_creates_timetable_for_one_class_only(self):
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY DAY",
            category="PRIMARY",
            study_days=["MON", "WED"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(8, 40),
        )
        profile.academic_levels.add(self.level)

        response = self.client.post(
            self.url,
            {"action": "generate_class", "class_id": str(self.academic_class.id)},
        )
        self.assertRedirects(response, self.url)
        lessons = list(GeneratedLearningLesson.objects.all())
        self.assertEqual(len(lessons), 2)
        self.assertTrue(all(lesson.academic_class_id == self.academic_class.id for lesson in lessons))

        page = self.client.get(self.url)
        self.assertContains(page, "Grade 1 East")
        self.assertContains(page, "Grade 1 West")
        self.assertContains(page, "No timetable generated yet")

    def test_reset_class_removes_lessons_for_that_class_only(self):
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY DAY",
            category="PRIMARY",
            study_days=["MON", "WED"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(8, 40),
        )
        profile.academic_levels.add(self.level)
        self.client.post(self.url, {"level_id": str(self.level.id)})
        self.assertEqual(GeneratedLearningLesson.objects.count(), 4)

        response = self.client.post(
            self.url,
            {"action": "reset_class", "class_id": str(self.academic_class.id)},
        )
        self.assertRedirects(response, self.url)
        remaining = list(GeneratedLearningLesson.objects.all())
        self.assertEqual(len(remaining), 2)
        self.assertTrue(all(lesson.academic_class_id == other_class.id for lesson in remaining))

    def test_manual_allocation_changes_subject_and_teacher_for_that_class_only(self):
        other_teacher = Employee.objects.create_user(
            employee_code="975310",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="TEACHER",
            email="beth.teacher@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        science = LearningArea.objects.create(name="Science", code="SCI")
        science.academic_levels.add(self.level)
        outsider = Employee.objects.create_user(
            employee_code="864209",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="CAROL",
            last_name="OUTSIDER",
            email="carol.outsider@example.com",
            phone_number="+254700000555",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        other_level = AcademicLevel.objects.create(
            name="Grade 2",
            code="G2",
            category="PRIMARY",
            order=2,
        )
        other_level_class = AcademicClass.objects.create(
            academic_level=other_level,
            name="Grade 2 East",
            code="G2E",
            order=1,
        )
        english = LearningArea.objects.create(name="English", code="ENG")
        english.academic_levels.add(other_level)
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=science,
            teacher=other_teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_level_class,
            learning_area=english,
            teacher=outsider,
        )
        generation = GeneratedLearningTimetable.objects.create(created_by=self.support)
        generation.academic_levels.add(self.level)
        east = GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        manual_url = reverse(
            "employees:it_support_timetable_page",
            kwargs={"tool": "manual-allocation"},
        )
        page = self.client.get(manual_url)
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "MATH")
        self.assertContains(page, "Change allocation")
        self.assertContains(page, '"available": true')
        self.assertContains(page, '"available": false')
        self.assertContains(page, "Grade 1 East")
        self.assertContains(page, "SCI")
        self.assertNotContains(page, "ENG")
        self.assertNotContains(page, "CAROL OUTSIDER")

        west = GeneratedLearningLesson.objects.get(academic_class=other_class)
        response = self.client.post(
            manual_url,
            {
                "lesson_id": str(east.id),
                "subject_id": str(science.id),
                "teacher_id": str(other_teacher.id),
            },
        )
        self.assertRedirects(response, manual_url)
        east.refresh_from_db()
        west.refresh_from_db()
        self.assertEqual(east.learning_area, science)
        self.assertEqual(east.teacher, other_teacher)
        self.assertEqual(west.learning_area, self.subject)
        self.assertEqual(west.teacher, self.teacher)
        page = self.client.get(manual_url)
        self.assertContains(page, "BETH TEACHER")
        self.assertContains(page, "ALI TEACHER")

    def test_manual_allocation_limits_options_to_that_class_only(self):
        other_teacher = Employee.objects.create_user(
            employee_code="975311",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="TEACHER",
            email="beth.teacher2@example.com",
            phone_number="+254700000446",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        science = LearningArea.objects.create(name="Science", code="SCI")
        science.academic_levels.add(self.level)
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=science,
            teacher=other_teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY DAY",
            category="PRIMARY",
            study_days=["MON"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(9, 20),
        )
        profile.academic_levels.add(self.level)
        generation = GeneratedLearningTimetable.objects.create(created_by=self.support)
        generation.academic_levels.add(self.level)
        GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        manual_url = reverse(
            "employees:it_support_timetable_page",
            kwargs={"tool": "manual-allocation"},
        )
        west_slot_key = f"{other_class.id}:MON:520"
        east_slot_key = f"{self.academic_class.id}:MON:520"
        page = self.client.get(manual_url)
        self.assertContains(page, f'"{east_slot_key}"')
        self.assertContains(page, f'"{west_slot_key}"')
        self.assertContains(page, '"subject_code": "SCI"')
        response = self.client.post(
            manual_url,
            {
                "slot_key": west_slot_key,
                "allocation": f"{science.id}:{other_teacher.id}",
            },
        )
        self.assertRedirects(response, manual_url)
        self.assertFalse(
            GeneratedLearningLesson.objects.filter(
                academic_class=other_class,
                learning_area=science,
            ).exists()
        )

    def test_manual_allocation_fills_free_slot_for_that_class_only(self):
        other_teacher = Employee.objects.create_user(
            employee_code="975312",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="TEACHER",
            email="beth.teacher3@example.com",
            phone_number="+254700000447",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        science = LearningArea.objects.create(name="Science", code="SCI")
        science.academic_levels.add(self.level)
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=science,
            teacher=other_teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY DAY",
            category="PRIMARY",
            study_days=["MON"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(9, 20),
        )
        profile.academic_levels.add(self.level)
        generation = GeneratedLearningTimetable.objects.create(created_by=self.support)
        generation.academic_levels.add(self.level)
        GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        manual_url = reverse(
            "employees:it_support_timetable_page",
            kwargs={"tool": "manual-allocation"},
        )
        slot_key = f"{self.academic_class.id}:MON:520"
        page = self.client.get(manual_url)
        self.assertContains(page, "Allocate")
        self.assertContains(page, f'"{slot_key}"')
        response = self.client.post(
            manual_url,
            {
                "slot_key": slot_key,
                "allocation": f"{science.id}:{other_teacher.id}",
            },
        )
        self.assertRedirects(response, manual_url)
        created = GeneratedLearningLesson.objects.get(
            academic_class=self.academic_class,
            weekday="MON",
            start_time=time(8, 40),
        )
        self.assertEqual(created.learning_area, science)
        self.assertEqual(created.teacher, other_teacher)
        page = self.client.get(manual_url)
        self.assertContains(page, "SCI")
        self.assertContains(page, "BETH TEACHER")
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        self.client.force_login(self.support)
        self.url = reverse(
            "employees:it_support_elearning_page",
            kwargs={"page": "timetable-generation"},
        )

    def test_page_loads_with_generate_controls(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Generate e-learning timetable")
        self.assertContains(response, 'data-open-modal="elearning-timetable-generation"')

    def test_fully_allocated_level_is_blocked_without_timetable_settings(self):
        ELearningSubjectAllocation.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        response = self.client.get(self.url)
        self.assertContains(response, "is-blocked")
        self.assertContains(response, "e-learning timetable settings")

    def test_fully_allocated_level_with_timetable_settings_can_be_generated(self):
        ELearningSubjectAllocation.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        profile = LearningScheduleProfile.objects.create(
            name="PRIMARY E-LEARNING SESSION",
            category="PRIMARY",
            kind=LearningScheduleProfile.Kind.ELEARNING,
            study_days=["MON", "WED"],
            lesson_duration_minutes=40,
            first_class_start_time=time(8, 0),
            last_class_end_time=time(8, 40),
        )
        profile.academic_levels.add(self.level)
        response = self.client.get(self.url)
        self.assertContains(response, "is-viable")
        self.assertContains(response, "PRIMARY E-LEARNING SESSION")

        response = self.client.post(self.url, {"level_id": str(self.level.id)})
        self.assertRedirects(response, self.url)
        generation = GeneratedELearningTimetable.objects.get()
        self.assertIn(self.level, generation.academic_levels.all())
        lessons = list(
            GeneratedELearningLesson.objects.filter(generation=generation).order_by("weekday")
        )
        self.assertEqual(len(lessons), 2)
        self.assertEqual({lesson.weekday for lesson in lessons}, {"MON", "WED"})
        self.assertEqual(lessons[0].start_time, time(8, 0))
        self.assertEqual(lessons[0].end_time, time(8, 40))
        page = self.client.get(self.url)
        self.assertContains(page, "MATH")
        self.assertContains(page, str(self.teacher.employment_number))
        self.assertContains(page, "Monday")
        self.assertContains(page, "Wednesday")

    def test_generate_query_opens_generation_popup(self):
        response = self.client.get(f"{self.url}?generate=1")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-modal="elearning-timetable-generation"')
        self.assertContains(response, "is-open")


class ExamSupervisorAllocationTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.accountant = Employee.objects.create_user(
            employee_code="975310",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="JOY",
            last_name="ACCOUNTS",
            email="accounts@example.com",
            phone_number="+254700000333",
            role=Employee.Role.ACCOUNTANT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.academic_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        self.client.force_login(self.support)

    def test_allocation_page_lists_level_subjects_and_teachers(self):
        response = self.client.get(
            reverse(
                "employees:it_support_exam_page",
                kwargs={"tool": "allocate-supervisors"},
            )
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Grade 1")
        self.assertContains(response, "Mathematics")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "Shuffle supervisors")
        self.assertContains(response, "classes merged")
        self.assertNotContains(response, "JOY ACCOUNTS")
        self.assertNotContains(response, 'name="supervisor_')

    def test_allocation_page_shows_class_subject_teacher(self):
        class_teacher = Employee.objects.create_user(
            employee_code="246800",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="CLASS",
            email="class.teacher@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=class_teacher,
        )
        response = self.client.get(
            reverse(
                "employees:it_support_exam_page",
                kwargs={"tool": "allocate-supervisors"},
            )
        )
        self.assertContains(response, "Class teachers")
        self.assertContains(response, "Supervisor")
        self.assertContains(response, "BETH CLASS")

    def test_shuffle_assigns_one_supervisor_to_all_classes_in_the_level(self):
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        english = LearningArea.objects.create(name="English", code="ENG")
        english.academic_levels.add(self.level)
        other_teacher = Employee.objects.create_user(
            employee_code="246800",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="CLASS",
            email="class.teacher@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=english,
            teacher=other_teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=other_class,
            learning_area=english,
            teacher=other_teacher,
        )
        response = self.client.post(
            reverse(
                "employees:it_support_exam_page",
                kwargs={"tool": "allocate-supervisors"},
            ),
            {"level_id": str(self.level.id)},
        )
        self.assertRedirects(
            response,
            reverse(
                "employees:it_support_exam_page",
                kwargs={"tool": "allocate-supervisors"},
            ),
        )
        math_supervisors = {
            item.supervisor_id
            for item in ExamSupervisorAllocation.objects.filter(learning_area=self.subject)
        }
        english_supervisors = {
            item.supervisor_id
            for item in ExamSupervisorAllocation.objects.filter(learning_area=english)
        }
        self.assertEqual(len(math_supervisors), 1)
        self.assertEqual(len(english_supervisors), 1)
        self.assertEqual(
            ExamSupervisorAllocation.objects.filter(learning_area=self.subject).count(),
            2,
        )
        self.assertNotEqual(next(iter(math_supervisors)), self.teacher.id)
        self.assertNotEqual(next(iter(english_supervisors)), other_teacher.id)
        self.assertNotEqual(math_supervisors, english_supervisors)


class ExamTimetableGenerationTests(TestCase):
    def setUp(self):
        self.support = Employee.objects.create_user(
            employee_code="246810",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="KIM",
            last_name="ITOTE",
            email="it.support@example.com",
            phone_number="+254700000111",
            role=Employee.Role.IT_SUPPORT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.academic_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        self.year = AcademicYear.objects.create(
            name="2026",
            start_date=date(2026, 1, 5),
            end_date=date(2026, 11, 20),
            is_current=True,
        )
        self.term = AcademicTerm.objects.create(
            academic_year=self.year,
            name="TERM 3",
            start_date=date(2026, 8, 10),
            end_date=date(2026, 11, 20),
            opening_date=date(2026, 8, 12),
            midterm_date=date(2026, 9, 20),
            closing_date=date(2026, 11, 14),
            order=3,
        )
        self.client.force_login(self.support)
        self.url = reverse(
            "employees:it_support_exam_page",
            kwargs={"tool": "exam-timetable-generation"},
        )

    def _exam_profile(self):
        profile = ExamScheduleProfile.objects.create(
            name="PRIMARY EXAM SESSION",
            category="PRIMARY",
            first_exam_start_time=time(8, 0),
            last_exam_end_time=time(10, 0),
            exam_session_duration_minutes=120,
        )
        profile.academic_levels.add(self.level)
        return profile

    def _generate_post(self, **overrides):
        payload = {
            "level_id": str(self.level.id),
            "academic_year_id": str(self.year.id),
            "academic_term_id": str(self.term.id),
            "exam_name": "END EXAM",
            "exam_start_date": "2026-08-19",
        }
        payload.update(overrides)
        return self.client.post(self.url, payload)

    def test_generation_page_opens_popup_and_blocks_levels_without_exam_profile(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-open-modal="exam-timetable-generation"')
        self.assertContains(response, "Grade 1")
        self.assertContains(response, "is-blocked")
        self.assertContains(response, "Assessment starts on")
        self.assertContains(response, "Assessment ends on")
        self.assertContains(response, "Assessment name")
        self.assertContains(response, 'name="exam_name"')
        self.assertContains(response, 'name="academic_year_id"')
        self.assertContains(response, "uppercase-input")
        self.assertNotContains(response, "Exam dates until")

    def test_level_without_supervisors_can_be_generated_with_exam_profile(self):
        self._exam_profile()
        response = self.client.get(self.url)
        self.assertContains(response, "is-viable")
        self.assertContains(response, "without a supervisor")

        response = self._generate_post()
        self.assertRedirects(response, self.url)
        generation = GeneratedExamTimetable.objects.get()
        sitting = GeneratedExamSitting.objects.get(generation=generation)
        self.assertIsNone(sitting.supervisor_id)
        self.assertEqual(sitting.learning_area, self.subject)

    def test_generate_query_opens_the_generation_popup(self):
        response = self.client.get(self.url, {"generate": "1"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="exam-timetable-generation-modal"')
        self.assertContains(response, "app-modal is-open")

    def test_fully_allocated_level_is_blocked_without_exam_profile(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        response = self.client.get(self.url)
        self.assertContains(response, "is-blocked")
        self.assertContains(response, "assessment timetable settings")

    def test_fully_allocated_level_with_exam_profile_can_be_generated(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        response = self.client.get(self.url)
        self.assertContains(response, "is-viable")
        self.assertContains(response, "PRIMARY EXAM SESSION")

        response = self._generate_post()
        self.assertRedirects(response, self.url)
        generation = GeneratedExamTimetable.objects.get()
        self.assertIn(self.level, generation.academic_levels.all())
        self.assertEqual(generation.name, "END EXAM")
        self.assertEqual(generation.academic_year, self.year)
        self.assertEqual(generation.academic_term, self.term)
        self.assertEqual(generation.start_date, date(2026, 8, 19))
        self.assertEqual(generation.end_date, date(2026, 8, 19))
        sittings = list(GeneratedExamSitting.objects.filter(generation=generation))
        self.assertEqual(len(sittings), 1)
        self.assertEqual(sittings[0].weekday, "WED")
        self.assertEqual(sittings[0].exam_date, date(2026, 8, 19))
        self.assertEqual(sittings[0].start_time, time(8, 0))
        self.assertEqual(sittings[0].end_time, time(10, 0))
        page = self.client.get(self.url)
        self.assertContains(page, "Grade 1 East")
        self.assertContains(page, "END EXAM")
        self.assertContains(page, "MATH")
        self.assertContains(page, "ALI TEACHER")
        self.assertContains(page, "Wednesday 19 Aug 2026")
        self.assertContains(page, "TERM 3")

    def test_exam_name_is_required_and_stored_uppercase(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        response = self._generate_post(exam_name="")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Enter a name for this assessment.")
        self.assertFalse(GeneratedExamTimetable.objects.exists())

        response = self._generate_post(exam_name="end exam")
        self.assertRedirects(response, self.url)
        generation = GeneratedExamTimetable.objects.get()
        self.assertEqual(generation.name, "END EXAM")
        self.assertEqual(generation.display_name, "END EXAM")

    def test_academic_year_must_be_a_registered_year(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        other_year = AcademicYear.objects.create(
            name="2025",
            start_date=date(2025, 1, 5),
            end_date=date(2025, 11, 20),
            is_current=False,
        )
        other_term = AcademicTerm.objects.create(
            academic_year=other_year,
            name="TERM 1",
            start_date=date(2025, 1, 5),
            end_date=date(2025, 4, 20),
            opening_date=date(2025, 1, 6),
            midterm_date=date(2025, 2, 20),
            closing_date=date(2025, 4, 15),
            order=1,
        )
        response = self._generate_post(academic_year_id="999999")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select a registered academic year.")
        self.assertFalse(GeneratedExamTimetable.objects.exists())

        response = self._generate_post(
            academic_year_id=str(other_year.id),
            academic_term_id=str(self.term.id),
            exam_start_date="2025-02-01",
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select an academic term from the selected academic year.")
        self.assertFalse(GeneratedExamTimetable.objects.exists())

        response = self._generate_post(
            academic_year_id=str(other_year.id),
            academic_term_id=str(other_term.id),
            exam_start_date="2025-02-01",
        )
        self.assertRedirects(response, self.url)
        generation = GeneratedExamTimetable.objects.get()
        self.assertEqual(generation.academic_year, other_year)
        self.assertEqual(generation.academic_term, other_term)

    def test_exam_dates_must_fall_inside_the_selected_term(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        response = self._generate_post(exam_start_date="2026-07-01")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "assessment start date must fall inside")
        self.assertFalse(GeneratedExamSitting.objects.exists())

    def test_subjects_are_placed_once_and_end_date_is_automatic(self):
        english = LearningArea.objects.create(name="English", code="ENG")
        science = LearningArea.objects.create(name="Science", code="SCI")
        english.academic_levels.add(self.level)
        science.academic_levels.add(self.level)
        for subject in (self.subject, english, science):
            ExamSupervisorAllocation.objects.create(
                academic_class=self.academic_class,
                learning_area=subject,
                supervisor=self.teacher,
            )
        ExamScheduleProfile.objects.create(
            name="PRIMARY EXAM SESSION",
            category="PRIMARY",
            first_exam_start_time=time(8, 0),
            last_exam_end_time=time(12, 0),
            exam_session_duration_minutes=120,
        ).academic_levels.add(self.level)

        response = self._generate_post()
        self.assertRedirects(response, self.url)
        generation = GeneratedExamTimetable.objects.get()
        self.assertEqual(generation.start_date, date(2026, 8, 19))
        self.assertEqual(generation.end_date, date(2026, 8, 20))
        sittings = list(
            GeneratedExamSitting.objects.filter(generation=generation).order_by("exam_date", "start_time")
        )
        self.assertEqual(len(sittings), 3)
        self.assertEqual({sitting.learning_area.code for sitting in sittings}, {"MATH", "ENG", "SCI"})
        self.assertEqual({sitting.exam_date for sitting in sittings}, {date(2026, 8, 19), date(2026, 8, 20)})
        self.assertEqual(len([item for item in sittings if item.exam_date == date(2026, 8, 19)]), 2)
        self.assertEqual(len([item for item in sittings if item.exam_date == date(2026, 8, 20)]), 1)

    def test_generation_leaves_blank_when_the_same_supervisor_would_collide(self):
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        ExamSupervisorAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()

        response = self._generate_post()
        self.assertRedirects(response, self.url)
        sittings = list(GeneratedExamSitting.objects.all())
        self.assertEqual(len(sittings), 2)
        self.assertEqual({sitting.supervisor_id for sitting in sittings}, {self.teacher.id})
        page = self.client.get(self.url)
        self.assertContains(page, "MATH")
        self.assertNotContains(page, "ALI TEACHER")

    def test_manual_allocation_changes_supervisor_and_lists_busy_teachers(self):
        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        other_teacher = Employee.objects.create_user(
            employee_code="246813",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="BETH",
            last_name="TEACHER",
            email="beth.teacher@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        ExamSupervisorAllocation.objects.create(
            academic_class=other_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        west = GeneratedExamSitting.objects.get(academic_class=other_class)
        manual_url = reverse("employees:exam_manual_supervisor_allocation")
        page = self.client.get(manual_url)
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "MATH")
        self.assertContains(page, "Change supervisor")
        self.assertContains(page, '"available": true')
        self.assertContains(page, '"available": false')
        self.assertContains(page, "Grade 1 East")

        response = self.client.post(
            manual_url,
            {"sitting_id": str(west.id), "supervisor_id": str(other_teacher.id)},
        )
        self.assertRedirects(response, manual_url)
        west.refresh_from_db()
        self.assertEqual(west.supervisor, other_teacher)
        page = self.client.get(manual_url)
        self.assertContains(page, "BETH TEACHER")
        self.assertContains(page, "ALI TEACHER")

    def test_exam_records_lists_generated_exams(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()

        response = self.client.get(
            reverse("employees:it_support_exam_page", kwargs={"tool": "exam-records"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Registered assessments")
        self.assertContains(response, "END EXAM")
        self.assertContains(response, "TERM 3")
        self.assertContains(response, "Grade 1")
        self.assertContains(response, "19 Aug 2026")
        self.assertContains(response, "Open")
        self.assertNotContains(response, ">Edit<")
        self.assertNotContains(response, ">Delete<")
        self.assertNotContains(response, "No assessments registered yet")
        generation = GeneratedExamTimetable.objects.get()
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": generation.id})
        self.assertContains(response, detail_url)

        page = self.client.get(detail_url)
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "END EXAM")
        self.assertContains(page, "Academic levels")
        self.assertContains(page, "Grade 1")
        self.assertContains(page, "Back to all assessments")
        self.assertContains(page, "Open marks")
        self.assertContains(page, 'data-open-modal="exam-edit"')
        self.assertContains(page, 'data-open-modal="exam-status"')
        self.assertContains(page, 'data-open-modal="exam-deadline"')
        self.assertContains(page, 'data-open-modal="exam-delete"')
        self.assertNotContains(page, "Select an academic level")
        level_url = reverse(
            "employees:exam_record_level",
            kwargs={"exam_id": generation.id, "level_id": self.level.id},
        )
        self.assertContains(page, level_url)

        from apps.admissions.models import ParentGuardian, Student

        other_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        parent = ParentGuardian.objects.create(
            full_name="JANE DOE",
            relationship_to_student="MOTHER",
            phone_number="+254700000333",
            email="jane.exam@example.com",
        )
        Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="G1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        Student.objects.create(
            first_name="BEN",
            last_name="WEST",
            date_of_birth="2018-02-02",
            gender=Student.Gender.MALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1002",
            class_group="G1W",
            assessment_number="A1002",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )

        level_page = self.client.get(level_url)
        self.assertEqual(level_page.status_code, 200)
        self.assertContains(level_page, "ANN EAST")
        self.assertContains(level_page, "BEN WEST")
        self.assertContains(level_page, "1001")
        self.assertContains(level_page, "1002")
        self.assertContains(level_page, "MATH")
        self.assertContains(level_page, "Edit")
        self.assertContains(level_page, "Save marks")
        self.assertContains(level_page, "readonly")
        self.assertNotContains(level_page, "/ 100 → 100%")
        self.assertNotContains(level_page, "Assessment")
        self.assertNotContains(level_page, "Gender")
        self.assertNotContains(level_page, "A1001")
        self.assertContains(level_page, "All classes")
        self.assertContains(level_page, "Grade 1 East")
        self.assertContains(level_page, "Grade 1 West")

        filtered = self.client.get(level_url, {"class_id": str(self.academic_class.id)})
        self.assertContains(filtered, "ANN EAST")
        self.assertNotContains(filtered, "BEN WEST")

        west_filtered = self.client.get(level_url, {"class_id": str(other_class.id)})
        self.assertContains(west_filtered, "BEN WEST")
        self.assertNotContains(west_filtered, "ANN EAST")

        ann = Student.objects.get(admission_number="1001")
        save_response = self.client.post(
            level_url,
            {f"mark_{ann.id}_{self.subject.id}": "78"},
        )
        self.assertRedirects(save_response, level_url)
        mark = ExamMark.objects.get(student=ann, learning_area=self.subject)
        self.assertEqual(mark.marks, 78)
        saved_page = self.client.get(level_url)
        self.assertContains(saved_page, 'value="78"')
        self.assertContains(saved_page, "Student marks were saved.")

        invalid = self.client.post(
            level_url,
            {f"mark_{ann.id}_{self.subject.id}": "140"},
        )
        self.assertEqual(invalid.status_code, 200)
        self.assertContains(invalid, "from 0 to 100")
        mark.refresh_from_db()
        self.assertEqual(mark.marks, 78)

        ExamSubjectSetting.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            out_of_marks=50,
        )
        converted = self.client.post(
            level_url,
            {f"mark_{ann.id}_{self.subject.id}": "50"},
        )
        self.assertRedirects(converted, level_url)
        mark.refresh_from_db()
        self.assertEqual(mark.marks, 25)
        self.assertEqual(mark.out_of_marks, 50)
        converted_page = self.client.get(level_url)
        self.assertContains(converted_page, 'value="50"')
        self.assertNotContains(converted_page, "/ 50 → 100%")

    def test_exam_record_level_shows_combined_subject_results(self):
        from apps.admissions.models import ParentGuardian, Student

        art = LearningArea.objects.create(name="Art", code="ART")
        art.academic_levels.add(self.level)
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=art,
            supervisor=self.teacher,
        )
        math_setting = ExamSubjectSetting.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            out_of_marks=50,
        )
        art_setting = ExamSubjectSetting.objects.create(
            academic_level=self.level,
            learning_area=art,
            out_of_marks=50,
        )
        combined = CombinedExamSubject.objects.create(
            academic_level=self.level,
            name="CREATIVE ARTS",
            code="CA-COMB",
        )
        CombinedExamSubjectComponent.objects.bulk_create(
            [
                CombinedExamSubjectComponent(
                    combined_subject=combined,
                    subject_setting=math_setting,
                    position=1,
                ),
                CombinedExamSubjectComponent(
                    combined_subject=combined,
                    subject_setting=art_setting,
                    position=2,
                ),
            ]
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        parent = ParentGuardian.objects.create(
            full_name="JANE DOE",
            relationship_to_student="MOTHER",
            phone_number="+254700000333",
            email="jane.exam@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="G1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=generation,
            student=student,
            learning_area=self.subject,
            marks=40,
            out_of_marks=50,
        )
        ExamMark.objects.create(
            generation=generation,
            student=student,
            learning_area=art,
            marks=30,
            out_of_marks=50,
        )
        level_url = reverse(
            "employees:exam_record_level",
            kwargs={"exam_id": generation.id, "level_id": self.level.id},
        )
        response = self.client.get(level_url, {"class_id": str(self.academic_class.id)})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "CA-COMB")
        self.assertContains(response, "MATH + ART")
        self.assertNotContains(response, 'name="mark_')
        self.assertContains(response, ">70<")

    def test_exam_record_can_be_updated(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": generation.id})
        response = self.client.post(
            reverse("employees:update_exam_record", kwargs={"exam_id": generation.id}),
            {
                "exam_name": "midterm exam",
                "academic_year_id": str(self.year.id),
                "academic_term_id": str(self.term.id),
                "start_date": "2026-08-12",
                "end_date": "2026-08-14",
                "academic_levels": [str(self.level.id)],
                "next": detail_url,
            },
        )
        self.assertRedirects(response, detail_url)
        generation.refresh_from_db()
        self.assertEqual(generation.name, "MIDTERM EXAM")
        self.assertEqual(generation.start_date, date(2026, 8, 12))
        self.assertEqual(generation.end_date, date(2026, 8, 14))
        self.assertEqual(list(generation.academic_levels.values_list("id", flat=True)), [self.level.id])
        page = self.client.get(detail_url)
        self.assertContains(page, "MIDTERM EXAM")
        self.assertContains(page, "was updated")

    def test_exam_record_status_and_deadline_can_be_updated(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": generation.id})
        self.assertEqual(generation.status, GeneratedExamTimetable.Status.IN_SESSION)

        status_response = self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": generation.id}),
            {"status": GeneratedExamTimetable.Status.MARKING, "next": detail_url},
        )
        self.assertRedirects(status_response, detail_url)
        generation.refresh_from_db()
        self.assertEqual(generation.status, GeneratedExamTimetable.Status.MARKING)

        deadline_response = self.client.post(
            reverse("employees:update_exam_record_deadline", kwargs={"exam_id": generation.id}),
            {"deadline": "2026-08-20T17:30", "next": detail_url},
        )
        self.assertRedirects(deadline_response, detail_url)
        generation.refresh_from_db()
        self.assertIsNotNone(generation.deadline)

    def test_current_exam_toggle_stays_on_when_status_advances_to_analysing(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": generation.id})

        self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": generation.id}),
            {"status": GeneratedExamTimetable.Status.MARKING, "next": detail_url},
        )
        self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": generation.id}),
            {"status": GeneratedExamTimetable.Status.ANALYSING, "next": detail_url},
        )

        generation.refresh_from_db()
        self.assertEqual(generation.status, GeneratedExamTimetable.Status.ANALYSING)
        page = self.client.get(detail_url)
        self.assertContains(page, 'data-exam-current-value" value="1"')

    def test_current_exam_toggle_stays_on_when_status_advances_to_marking(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": generation.id})

        self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": generation.id}),
            {"status": GeneratedExamTimetable.Status.MARKING, "next": detail_url},
        )

        page = self.client.get(detail_url)
        self.assertContains(page, "data-exam-current-input")
        self.assertContains(page, 'data-exam-current-value" value="1"')

    def test_current_exam_toggle_turns_off_when_exam_is_published(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": generation.id})

        self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": generation.id}),
            {"status": GeneratedExamTimetable.Status.MARKING, "next": detail_url},
        )
        self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": generation.id}),
            {"status": GeneratedExamTimetable.Status.PUBLISHED, "next": detail_url},
        )

        page = self.client.get(detail_url)
        self.assertNotContains(page, "data-exam-current-input")
        self.assertContains(page, "Published exams cannot be set as current")

    def test_only_current_exam_can_change_status(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        current_exam = GeneratedExamTimetable.objects.get()
        scheduled_exam = GeneratedExamTimetable.objects.create(
            name="NEXT EXAM",
            academic_year=self.year,
            academic_term=self.term,
            start_date=date(2026, 9, 1),
            end_date=date(2026, 9, 3),
            status=GeneratedExamTimetable.Status.SCHEDULED,
        )
        scheduled_exam.academic_levels.add(self.level)
        scheduled_url = reverse("employees:exam_record_detail", kwargs={"exam_id": scheduled_exam.id})
        current_url = reverse("employees:exam_record_detail", kwargs={"exam_id": current_exam.id})

        blocked = self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": scheduled_exam.id}),
            {"status": GeneratedExamTimetable.Status.IN_SESSION, "next": scheduled_url},
        )
        self.assertRedirects(blocked, scheduled_url)
        scheduled_page = self.client.get(scheduled_url)
        self.assertContains(scheduled_page, "Only one assessment can be current at a time")
        self.assertNotContains(scheduled_page, 'data-open-modal="exam-status"')

        allowed = self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": current_exam.id}),
            {"status": GeneratedExamTimetable.Status.MARKING, "next": current_url},
        )
        self.assertRedirects(allowed, current_url)
        current_exam.refresh_from_db()
        self.assertEqual(current_exam.status, GeneratedExamTimetable.Status.MARKING)

    def test_second_generated_exam_is_scheduled_when_one_is_active(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post(exam_name="FIRST EXAM")
        first_exam = GeneratedExamTimetable.objects.get()
        self.assertEqual(first_exam.status, GeneratedExamTimetable.Status.IN_SESSION)

        self._generate_post(exam_name="SECOND EXAM", exam_start_date="2026-08-20")
        second_exam = GeneratedExamTimetable.objects.exclude(pk=first_exam.pk).get()
        self.assertEqual(second_exam.status, GeneratedExamTimetable.Status.SCHEDULED)

    def test_scheduled_exam_can_start_when_no_other_active_exam(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        current_exam = GeneratedExamTimetable.objects.get()
        current_url = reverse("employees:exam_record_detail", kwargs={"exam_id": current_exam.id})
        self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": current_exam.id}),
            {"status": GeneratedExamTimetable.Status.PUBLISHED, "next": current_url},
        )

        scheduled_exam = GeneratedExamTimetable.objects.create(
            name="NEXT EXAM",
            academic_year=self.year,
            academic_term=self.term,
            start_date=date(2026, 9, 1),
            end_date=date(2026, 9, 3),
            status=GeneratedExamTimetable.Status.SCHEDULED,
        )
        scheduled_exam.academic_levels.add(self.level)
        scheduled_url = reverse("employees:exam_record_detail", kwargs={"exam_id": scheduled_exam.id})

        response = self.client.post(
            reverse("employees:update_exam_record_status", kwargs={"exam_id": scheduled_exam.id}),
            {"status": GeneratedExamTimetable.Status.IN_SESSION, "next": scheduled_url},
        )
        self.assertRedirects(response, scheduled_url)
        scheduled_exam.refresh_from_db()
        self.assertEqual(scheduled_exam.status, GeneratedExamTimetable.Status.IN_SESSION)
        self.assertEqual(
            GeneratedExamTimetable.objects.filter(status=GeneratedExamTimetable.Status.IN_SESSION).count(),
            1,
        )

    def test_exam_records_list_shows_set_current_controls(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        records_url = reverse("employees:it_support_exam_page", kwargs={"tool": "exam-records"})
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": generation.id})
        response = self.client.get(records_url)
        self.assertNotContains(response, "Set current")
        self.assertNotContains(response, "Remove current")

        page = self.client.get(detail_url)
        self.assertContains(page, "Current assessment")
        self.assertContains(page, "data-exam-current-toggle-form")
        self.assertContains(page, "data-exam-current-input")

    def test_set_current_exam_from_detail_page(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        current_exam = GeneratedExamTimetable.objects.get()
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": current_exam.id})
        self.client.post(
            reverse("employees:set_current_exam_record", kwargs={"exam_id": current_exam.id}),
            {"is_current": "0", "next": detail_url},
        )
        current_exam.refresh_from_db()
        self.assertEqual(current_exam.status, GeneratedExamTimetable.Status.SCHEDULED)

        scheduled_exam = GeneratedExamTimetable.objects.create(
            name="NEXT EXAM",
            academic_year=self.year,
            academic_term=self.term,
            start_date=date(2026, 9, 1),
            end_date=date(2026, 9, 3),
            status=GeneratedExamTimetable.Status.SCHEDULED,
        )
        scheduled_exam.academic_levels.add(self.level)
        scheduled_detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": scheduled_exam.id})
        response = self.client.post(
            reverse("employees:set_current_exam_record", kwargs={"exam_id": scheduled_exam.id}),
            {"is_current": "1", "next": scheduled_detail_url},
        )
        self.assertRedirects(response, scheduled_detail_url)
        scheduled_exam.refresh_from_db()
        self.assertEqual(scheduled_exam.status, GeneratedExamTimetable.Status.IN_SESSION)

    def test_set_current_exam_from_list(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        current_exam = GeneratedExamTimetable.objects.get()
        records_url = reverse("employees:it_support_exam_page", kwargs={"tool": "exam-records"})
        self.client.post(
            reverse("employees:set_current_exam_record", kwargs={"exam_id": current_exam.id}),
            {"is_current": "0", "next": records_url},
        )
        current_exam.refresh_from_db()
        self.assertEqual(current_exam.status, GeneratedExamTimetable.Status.SCHEDULED)

        scheduled_exam = GeneratedExamTimetable.objects.create(
            name="NEXT EXAM",
            academic_year=self.year,
            academic_term=self.term,
            start_date=date(2026, 9, 1),
            end_date=date(2026, 9, 3),
            status=GeneratedExamTimetable.Status.SCHEDULED,
        )
        scheduled_exam.academic_levels.add(self.level)
        response = self.client.post(
            reverse("employees:set_current_exam_record", kwargs={"exam_id": scheduled_exam.id}),
            {"is_current": "1", "next": records_url},
        )
        self.assertRedirects(response, records_url)
        scheduled_exam.refresh_from_db()
        self.assertEqual(scheduled_exam.status, GeneratedExamTimetable.Status.IN_SESSION)
        page = self.client.get(records_url)
        self.assertContains(page, "NEXT EXAM")
        self.assertContains(page, "is now the current assessment")

    def test_set_current_switches_between_in_session_exams(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post(exam_name="FIRST EXAM")
        first_exam = GeneratedExamTimetable.objects.get()
        self._generate_post(exam_name="SECOND EXAM", exam_start_date="2026-08-20")
        second_exam = GeneratedExamTimetable.objects.exclude(pk=first_exam.pk).get()
        records_url = reverse("employees:it_support_exam_page", kwargs={"tool": "exam-records"})

        response = self.client.post(
            reverse("employees:set_current_exam_record", kwargs={"exam_id": second_exam.id}),
            {"is_current": "1", "next": records_url},
        )
        self.assertRedirects(response, records_url)
        first_exam.refresh_from_db()
        second_exam.refresh_from_db()
        self.assertEqual(first_exam.status, GeneratedExamTimetable.Status.SCHEDULED)
        self.assertEqual(second_exam.status, GeneratedExamTimetable.Status.IN_SESSION)

    def test_set_current_works_when_another_is_marking(self):
        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        current_exam = GeneratedExamTimetable.objects.get()
        current_exam.status = GeneratedExamTimetable.Status.MARKING
        current_exam.save(update_fields=["status"])
        scheduled_exam = GeneratedExamTimetable.objects.create(
            name="NEXT EXAM",
            academic_year=self.year,
            academic_term=self.term,
            start_date=date(2026, 9, 1),
            end_date=date(2026, 9, 3),
            status=GeneratedExamTimetable.Status.SCHEDULED,
        )
        scheduled_exam.academic_levels.add(self.level)
        records_url = reverse("employees:it_support_exam_page", kwargs={"tool": "exam-records"})
        response = self.client.post(
            reverse("employees:set_current_exam_record", kwargs={"exam_id": scheduled_exam.id}),
            {"is_current": "1", "next": records_url},
        )
        self.assertRedirects(response, records_url)
        current_exam.refresh_from_db()
        scheduled_exam.refresh_from_db()
        self.assertEqual(current_exam.status, GeneratedExamTimetable.Status.SCHEDULED)
        self.assertEqual(scheduled_exam.status, GeneratedExamTimetable.Status.IN_SESSION)
        detail_url = reverse("employees:exam_record_detail", kwargs={"exam_id": scheduled_exam.id})
        page = self.client.get(detail_url)
        self.assertContains(page, "Current assessment")
        self.assertContains(page, "data-exam-current-input")

    def test_exam_record_can_be_deleted(self):
        from apps.admissions.models import ParentGuardian, Student

        ExamSupervisorAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            supervisor=self.teacher,
        )
        self._exam_profile()
        self._generate_post()
        generation = GeneratedExamTimetable.objects.get()
        parent = ParentGuardian.objects.create(
            full_name="JANE DOE",
            relationship_to_student="MOTHER",
            phone_number="+254700000555",
            email="jane.delete@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="2001",
            class_group="G1E",
            assessment_number="A2001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=generation,
            student=student,
            learning_area=self.subject,
            marks=70,
        )
        self.assertTrue(GeneratedExamSitting.objects.filter(generation=generation).exists())
        self.assertTrue(ExamMark.objects.filter(generation=generation).exists())
        records_url = reverse("employees:it_support_exam_page", kwargs={"tool": "exam-records"})
        response = self.client.post(
            reverse("employees:delete_exam_record", kwargs={"exam_id": generation.id})
        )
        self.assertRedirects(response, records_url)
        self.assertFalse(GeneratedExamTimetable.objects.filter(pk=generation.id).exists())
        self.assertFalse(GeneratedExamSitting.objects.filter(generation_id=generation.id).exists())
        self.assertFalse(ExamMark.objects.filter(generation_id=generation.id).exists())
        page = self.client.get(records_url)
        self.assertContains(page, "was deleted")


class TeacherExamRecordsTests(TestCase):
    def setUp(self):
        self.teacher = Employee.objects.create_user(
            employee_code="135790",
            password="ReliablePass456",
            title=Employee.Title.MR,
            first_name="ALI",
            last_name="TEACHER",
            email="teacher@example.com",
            phone_number="+254700000222",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.other_teacher = Employee.objects.create_user(
            employee_code="246813",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="SARA",
            last_name="MWANGI",
            email="sara@example.com",
            phone_number="+254700000444",
            role=Employee.Role.TEACHER,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.accountant = Employee.objects.create_user(
            employee_code="111222",
            password="ReliablePass456",
            title=Employee.Title.MS,
            first_name="JOY",
            last_name="ACCOUNTS",
            email="accounts@example.com",
            phone_number="+254700000333",
            role=Employee.Role.ACCOUNTANT,
            approval_status=Employee.ApprovalStatus.APPROVED,
            is_active=True,
        )
        self.level = AcademicLevel.objects.create(
            name="Grade 1",
            code="G1",
            category="PRIMARY",
            order=1,
        )
        self.other_level = AcademicLevel.objects.create(
            name="Grade 2",
            code="G2",
            category="PRIMARY",
            order=2,
        )
        self.academic_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 East",
            code="G1E",
            order=1,
        )
        self.other_class = AcademicClass.objects.create(
            academic_level=self.other_level,
            name="Grade 2 West",
            code="G2W",
            order=1,
        )
        self.subject = LearningArea.objects.create(name="Mathematics", code="MATH")
        self.subject.academic_levels.add(self.level)
        other_subject = LearningArea.objects.create(name="English", code="ENG")
        other_subject.academic_levels.add(self.other_level)
        extra_subject = LearningArea.objects.create(name="English", code="ENG2")
        extra_subject.academic_levels.add(self.level)
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.academic_class,
            learning_area=extra_subject,
            teacher=self.other_teacher,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=self.other_class,
            learning_area=other_subject,
            teacher=self.other_teacher,
        )
        self.year = AcademicYear.objects.create(
            name="2026",
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_current=True,
        )
        self.term = AcademicTerm.objects.create(
            academic_year=self.year,
            name="TERM 3",
            start_date=date(2026, 8, 1),
            end_date=date(2026, 11, 30),
            opening_date=date(2026, 8, 1),
            midterm_date=date(2026, 9, 15),
            closing_date=date(2026, 11, 30),
            order=3,
        )
        self.exam = GeneratedExamTimetable.objects.create(
            academic_year=self.year,
            academic_term=self.term,
            start_date=date(2026, 8, 19),
            end_date=date(2026, 8, 21),
        )
        self.exam.academic_levels.add(self.level)
        self.client.force_login(self.teacher)

    def test_teacher_dashboard_links_to_exam_records(self):
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Assessment records")
        self.assertContains(response, reverse("employees:teacher_exam_records"))
        self.assertContains(response, "Subject attendance")
        self.assertContains(response, reverse("employees:teacher_subject_attendance"))
        self.assertContains(response, "Learning reports")
        self.assertContains(response, reverse("employees:teacher_learning_reports"))
        self.assertContains(response, "Session timetable")
        self.assertContains(response, "My class")
        self.assertContains(response, "E-learning")
        self.assertContains(response, "Locked")
        self.assertNotContains(response, reverse("employees:teacher_my_class"))
        self.assertNotContains(response, reverse("employees:teacher_elearning"))
        self.assertContains(
            response,
            "Your session timetable will appear here once IT Support generates",
        )

        learning_reports = self.client.get(reverse("employees:teacher_learning_reports"))
        self.assertEqual(learning_reports.status_code, 200)
        self.assertContains(learning_reports, "Learning reports")
        self.assertContains(learning_reports, "Generate learning report")
        self.assertContains(learning_reports, "Report date")
        self.assertContains(learning_reports, "Academic level")
        self.assertContains(learning_reports, "Class")
        self.assertContains(learning_reports, "Subject")
        self.assertContains(learning_reports, "Attendance")
        self.assertContains(learning_reports, "Lesson plan")
        self.assertContains(learning_reports, "Outcome")
        self.assertContains(learning_reports, "Generate report")
        self.assertContains(learning_reports, self.level.name)
        self.assertNotContains(learning_reports, "Print report")
        self.assertNotContains(learning_reports, "Open →")
        self.assertNotContains(learning_reports, "Attendance sessions")

        generated = self.client.get(
            reverse("employees:teacher_learning_reports"),
            {
                "generate": "1",
                "report_date": "2026-08-21",
                "report_type": "lesson_plan",
                "level_id": str(self.level.id),
                "class_id": str(self.academic_class.id),
                "subject_id": str(self.subject.id),
            },
        )
        self.assertEqual(generated.status_code, 200)
        self.assertContains(generated, "Print report")
        self.assertContains(generated, "Lesson plans")
        self.assertContains(generated, self.academic_class.name)
        self.assertContains(generated, self.subject.name)
        self.assertContains(generated, "21 Aug 2026")
        self.assertContains(generated, "elr-plan-card")
        self.assertNotContains(generated, "Open →")
        self.assertNotContains(generated, 'aria-label="Subject outcomes"')

        attendance_report = self.client.get(
            reverse("employees:teacher_learning_reports"),
            {
                "generate": "1",
                "report_date": "2026-08-21",
                "report_type": "attendance",
                "level_id": str(self.level.id),
                "class_id": str(self.academic_class.id),
                "subject_id": str(self.subject.id),
            },
        )
        self.assertEqual(attendance_report.status_code, 200)
        self.assertContains(attendance_report, "Attendance · 21 Aug 2026")
        self.assertContains(attendance_report, "Print report")

        outcome_report = self.client.get(
            reverse("employees:teacher_learning_reports"),
            {
                "generate": "1",
                "report_date": "2026-08-21",
                "report_type": "outcome",
                "level_id": str(self.level.id),
                "class_id": str(self.academic_class.id),
                "subject_id": str(self.subject.id),
            },
        )
        self.assertEqual(outcome_report.status_code, 200)
        self.assertContains(outcome_report, "Subject outcomes")
        self.assertContains(outcome_report, self.subject.name)

    def test_my_class_link_unlocks_when_teacher_is_class_teacher(self):
        self.academic_class.class_teacher = self.teacher
        self.academic_class.save(update_fields=["class_teacher"])
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "My class")
        self.assertContains(response, reverse("employees:teacher_my_class"))
        self.assertContains(response, "E-learning")
        self.assertNotContains(response, reverse("employees:teacher_elearning"))
        self.assertContains(response, "Locked")
        my_class = self.client.get(reverse("employees:teacher_my_class"))
        self.assertEqual(my_class.status_code, 200)
        self.assertContains(my_class, "Grade 1 East")
        self.assertContains(my_class, "Register class attendance")
        self.assertContains(my_class, "Students class attendance")
        self.assertContains(my_class, "Students discipline")
        self.assertContains(my_class, "Student books")
        self.assertContains(
            my_class,
            reverse(
                "employees:teacher_my_class_page",
                kwargs={"tool": "register-class-attendance"},
            ),
        )
        self.assertContains(
            my_class,
            reverse(
                "employees:teacher_my_class_page",
                kwargs={"tool": "students-class-attendance"},
            ),
        )
        self.assertNotContains(my_class, "Session timetable")
        self.assertNotContains(my_class, reverse("employees:teacher_subject_attendance"))
        self.assertNotContains(my_class, reverse("employees:teacher_exam_records"))

        register = self.client.get(
            reverse(
                "employees:teacher_my_class_page",
                kwargs={"tool": "register-class-attendance"},
            )
        )
        self.assertEqual(register.status_code, 200)
        self.assertContains(register, "Register class attendance")
        self.assertContains(register, "Students class attendance")
        self.assertContains(register, "Subject attendance")
        self.assertContains(register, "workspace-nav-label")
        self.assertContains(register, "Morning")
        self.assertContains(register, "Afternoon")
        self.assertContains(register, "Evening")
        self.assertContains(register, 'type="date"')
        self.assertContains(register, "Select date")
        self.assertContains(register, "Open calendar")
        self.assertContains(register, "Attendance day")
        self.assertContains(
            register,
            reverse(
                "employees:teacher_subject_attendance_class",
                kwargs={"class_id": self.academic_class.id},
            ),
        )
        self.assertNotContains(
            register,
            reverse(
                "employees:teacher_my_class_page",
                kwargs={"tool": "students-discipline"},
            ),
        )
        self.assertNotContains(
            register,
            reverse(
                "employees:teacher_my_class_page",
                kwargs={"tool": "student-books"},
            ),
        )

        analytics = self.client.get(
            reverse(
                "employees:teacher_my_class_page",
                kwargs={"tool": "students-class-attendance"},
            )
        )
        self.assertEqual(analytics.status_code, 200)
        self.assertContains(analytics, "Students class attendance")
        self.assertContains(analytics, "Attendance filter")
        self.assertContains(analytics, "Filter by")
        self.assertContains(analytics, "Morning")
        self.assertContains(analytics, "Afternoon")
        self.assertContains(analytics, "Evening")
        self.assertContains(analytics, "Register attendance")

    def test_class_teacher_can_save_morning_afternoon_evening_attendance(self):
        from apps.admissions.models import ParentGuardian, Student
        from apps.curriculum.models import ClassAttendanceRecord, ClassAttendanceSession

        self.academic_class.class_teacher = self.teacher
        self.academic_class.save(update_fields=["class_teacher"])
        parent = ParentGuardian.objects.create(
            full_name="PAT PARENT",
            relationship_to_student="MOTHER",
            phone_number="+254700009999",
            email="pat.parent@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="LEARNER",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="9001",
            class_group="Grade 1 East",
            assessment_number="A9001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        url = reverse(
            "employees:teacher_my_class_page",
            kwargs={"tool": "register-class-attendance"},
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ANN LEARNER")
        self.assertContains(response, f'name="morning_{student.id}"')
        self.assertContains(response, f'name="afternoon_{student.id}"')
        self.assertContains(response, f'name="evening_{student.id}"')

        save = self.client.post(
            url,
            {
                "class_id": str(self.academic_class.id),
                "attendance_date": "2026-08-21",
                "attendance_notes": "Assembly day",
                f"morning_{student.id}": "on",
                f"evening_{student.id}": "on",
            },
        )
        self.assertRedirects(
            save,
            f"{url}?class_id={self.academic_class.id}&date=2026-08-21",
        )
        session = ClassAttendanceSession.objects.get(
            academic_class=self.academic_class,
            attendance_date=date(2026, 8, 21),
        )
        self.assertEqual(session.notes, "Assembly day")
        record = ClassAttendanceRecord.objects.get(session=session, student=student)
        self.assertTrue(record.morning)
        self.assertFalse(record.afternoon)
        self.assertTrue(record.evening)

    def test_my_class_page_redirects_when_teacher_is_not_class_teacher(self):
        response = self.client.get(reverse("employees:teacher_my_class"))
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "teacher"}),
        )

    def test_class_teacher_subject_attendance_lists_students_and_subject_status(self):
        from apps.admissions.models import ParentGuardian, Student
        from apps.curriculum.models import SubjectAttendanceRecord, SubjectAttendanceSession

        self.academic_class.class_teacher = self.teacher
        self.academic_class.save(update_fields=["class_teacher"])
        parent = ParentGuardian.objects.create(
            full_name="PAT PARENT",
            relationship_to_student="MOTHER",
            phone_number="+254700009901",
            email="pat.matrix@example.com",
        )
        student = Student.objects.create(
            first_name="CARA",
            last_name="LEARNER",
            date_of_birth="2018-03-03",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="9002",
            class_group="Grade 1 East",
            assessment_number="A9002",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        session = SubjectAttendanceSession.objects.create(
            allocation=ClassSubjectAllocation.objects.get(
                academic_class=self.academic_class,
                learning_area=self.subject,
            ),
            lesson_date=date(2026, 8, 21),
            taken_by=self.teacher,
        )
        SubjectAttendanceRecord.objects.create(
            session=session,
            student=student,
            status=SubjectAttendanceRecord.Status.PRESENT,
        )
        response = self.client.get(
            reverse(
                "employees:teacher_subject_attendance_class",
                kwargs={"class_id": self.academic_class.id},
            )
            + "?date=2026-08-21"
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "CARA LEARNER")
        self.assertContains(response, "MATH")
        self.assertContains(response, "Present")
        self.assertContains(response, "Attendance filter")
        self.assertContains(response, "Filter by")
        self.assertContains(response, 'name="scope"')
        self.assertContains(response, "Day")
        self.assertContains(response, "Period")
        self.assertContains(response, "Term")
        self.assertContains(response, "Academic year")
        self.assertNotContains(response, "Your subjects")

        term_view = self.client.get(
            reverse(
                "employees:teacher_subject_attendance_class",
                kwargs={"class_id": self.academic_class.id},
            )
            + f"?scope=term&year={self.year.id}&term={self.term.id}"
        )
        self.assertEqual(term_view.status_code, 200)
        self.assertContains(term_view, "TERM 3")
        self.assertContains(term_view, "100%")
        self.assertContains(term_view, "CARA")

        period_view = self.client.get(
            reverse(
                "employees:teacher_subject_attendance_class",
                kwargs={"class_id": self.academic_class.id},
            )
            + f"?scope=period&year={self.year.id}&term={self.term.id}&period=1"
        )
        self.assertEqual(period_view.status_code, 200)
        self.assertContains(period_view, "Period 1")
        self.assertContains(period_view, "100%")

        year_view = self.client.get(
            reverse(
                "employees:teacher_subject_attendance_class",
                kwargs={"class_id": self.academic_class.id},
            )
            + f"?scope=year&year={self.year.id}"
        )
        self.assertEqual(year_view.status_code, 200)
        self.assertContains(year_view, "Academic year 2026")
        self.assertContains(year_view, "100%")

    def test_subject_attendance_class_opens_subjects_and_profile(self):
        response = self.client.get(reverse("employees:teacher_subject_attendance"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Grade 1 East")
        self.assertNotContains(response, "Session timetable")
        self.assertNotContains(response, reverse("employees:teacher_exam_records"))
        self.assertNotContains(response, "workspace-nav-label")
        self.assertContains(
            response,
            reverse(
                "employees:teacher_subject_attendance_class",
                kwargs={"class_id": self.academic_class.id},
            ),
        )
        class_page = self.client.get(
            reverse(
                "employees:teacher_subject_attendance_class",
                kwargs={"class_id": self.academic_class.id},
            )
        )
        self.assertEqual(class_page.status_code, 200)
        self.assertContains(class_page, "Mathematics")
        self.assertContains(class_page, "Your subjects")
        profile_url = reverse(
            "employees:teacher_subject_attendance_profile",
            kwargs={
                "class_id": self.academic_class.id,
                "subject_id": self.subject.id,
            },
        )
        self.assertContains(class_page, profile_url)
        profile = self.client.get(profile_url)
        self.assertEqual(profile.status_code, 200)
        self.assertContains(profile, "Lesson plan")
        self.assertContains(profile, "Attendance")
        self.assertContains(profile, "Class subject outcome")
        save_plan = self.client.post(
            profile_url,
            {
                "form_action": "lesson_plan",
                "strand": "Numbers",
                "substrand": "Fractions",
                "lesson_learning_outcomes": "Identify halves",
                "key_inquiry_questions": "What is a half?",
                "core_competencies": "Critical thinking",
                "values": "Responsibility",
                "pcis": "Health education",
                "learning_resources": "Counters",
                "organization_of_learning": "Group work",
                "introduction": "Review wholes",
                "lesson_development": "Fold paper into halves",
            },
        )
        self.assertRedirects(save_plan, f"{profile_url}#lesson-plan")
        self.assertTrue(
            ClassSubjectLessonPlan.objects.filter(
                allocation__academic_class=self.academic_class,
                allocation__learning_area=self.subject,
                strand="Numbers",
                substrand="Fractions",
            ).exists()
        )

    def test_teacher_dashboard_shows_session_timetable(self):
        generation = GeneratedLearningTimetable.objects.create()
        generation.academic_levels.add(self.level)
        GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            academic_class=self.academic_class,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="MON",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        GeneratedLearningLesson.objects.create(
            generation=generation,
            academic_level=self.other_level,
            academic_class=self.other_class,
            learning_area=LearningArea.objects.get(code="ENG"),
            teacher=self.other_teacher,
            weekday="TUE",
            period_name="Period 1",
            start_time=time(8, 0),
            end_time=time(8, 40),
        )
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "This week")
        self.assertContains(response, "Learning this week")
        self.assertContains(response, "1 lesson")
        self.assertContains(response, "MATH")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "Period 1")
        self.assertNotContains(response, "ENG")
        self.assertNotContains(response, "Grade 2 West")
        self.assertNotContains(
            response,
            "Your session timetable will appear here once IT Support generates",
        )

    def test_elearning_link_unlocks_and_shows_timetable_when_allocated(self):
        ELearningSubjectAllocation.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        generation = GeneratedELearningTimetable.objects.create()
        generation.academic_levels.add(self.level)
        GeneratedELearningLesson.objects.create(
            generation=generation,
            academic_level=self.level,
            learning_area=self.subject,
            teacher=self.teacher,
            weekday="WED",
            period_name="Session 1",
            start_time=time(9, 0),
            end_time=time(9, 40),
        )
        response = self.client.get(
            reverse("employees:role_dashboard", kwargs={"role": "teacher"})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("employees:teacher_elearning"))
        self.assertContains(response, "E-learning this week")
        self.assertContains(response, "1 session")
        self.assertContains(response, "MATH")
        self.assertContains(response, "Session 1")
        self.assertContains(response, "Wednesday")
        page = self.client.get(reverse("employees:teacher_elearning"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Allocated subjects")
        self.assertContains(page, "MATH")
        self.assertContains(page, "E-learning this week")
        self.assertContains(page, "E-learning attendance")
        self.assertContains(page, "E-learning assessments")
        self.assertContains(page, "Learning materials")
        self.assertContains(page, "E-learning reports")
        self.assertNotContains(page, "Attendance reports")
        self.assertNotContains(page, "Lesson plan reports")
        self.assertNotContains(page, "Outcome reports")
        self.assertContains(page, reverse("employees:teacher_elearning_page", kwargs={"tool": "attendance"}))
        self.assertContains(
            page,
            reverse("employees:teacher_elearning_page", kwargs={"tool": "e-learning-reports"}),
        )
        attendance = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "attendance"})
        )
        self.assertEqual(attendance.status_code, 200)
        self.assertContains(attendance, "E-learning attendance")
        self.assertContains(attendance, "Open subjects")
        self.assertContains(attendance, self.level.name)
        self.assertContains(attendance, "E-learning reports")
        self.assertNotContains(attendance, "Attendance reports")
        self.assertNotContains(attendance, "Lesson plan reports")
        self.assertNotContains(attendance, "Outcome reports")
        self.assertNotContains(attendance, "Tools coming soon")
        self.assertNotContains(attendance, reverse("employees:teacher_subject_attendance"))

        elearning_report = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "e-learning-reports"})
        )
        self.assertEqual(elearning_report.status_code, 200)
        self.assertContains(elearning_report, "E-learning reports")
        self.assertContains(elearning_report, "Generate e-learning report")
        self.assertContains(elearning_report, "Report date")
        self.assertContains(elearning_report, "Academic level")
        self.assertContains(elearning_report, "Subject")
        self.assertContains(elearning_report, "Attendance")
        self.assertContains(elearning_report, "Lesson plan")
        self.assertContains(elearning_report, "Outcome")
        self.assertContains(elearning_report, "Generate report")
        self.assertContains(elearning_report, self.level.name)
        self.assertNotContains(elearning_report, "Print report")
        self.assertNotContains(elearning_report, "Tools coming soon")
        self.assertNotContains(elearning_report, "Attendance reports")
        self.assertNotContains(elearning_report, "Lesson plan reports")
        self.assertNotContains(elearning_report, "Outcome reports")
        self.assertNotContains(elearning_report, "Open →")

        generated = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "e-learning-reports"}),
            {
                "generate": "1",
                "report_date": "2026-08-21",
                "report_type": "lesson_plan",
                "level_id": str(self.level.id),
                "subject_id": str(self.subject.id),
            },
        )
        self.assertEqual(generated.status_code, 200)
        self.assertContains(generated, "Print report")
        self.assertContains(generated, "Lesson plans")
        self.assertContains(generated, self.level.name)
        self.assertContains(generated, self.subject.name)
        self.assertContains(generated, "21 Aug 2026")
        self.assertContains(generated, "elr-plan-card")
        self.assertNotContains(generated, "Attendance ·")
        self.assertNotContains(generated, 'aria-label="Subject outcomes"')
        self.assertNotContains(
            generated,
            reverse(
                "employees:teacher_elearning_attendance_profile",
                kwargs={"level_id": self.level.id, "subject_id": self.subject.id},
            ),
        )

        attendance_report = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "e-learning-reports"}),
            {
                "generate": "1",
                "report_date": "2026-08-21",
                "report_type": "attendance",
                "level_id": str(self.level.id),
                "subject_id": str(self.subject.id),
            },
        )
        self.assertEqual(attendance_report.status_code, 200)
        self.assertContains(attendance_report, "Attendance · 21 Aug 2026")
        self.assertContains(attendance_report, "Print report")

        outcome_report = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "e-learning-reports"}),
            {
                "generate": "1",
                "report_date": "2026-08-21",
                "report_type": "outcome",
                "level_id": str(self.level.id),
                "subject_id": str(self.subject.id),
            },
        )
        self.assertEqual(outcome_report.status_code, 200)
        self.assertContains(outcome_report, "Subject outcomes")
        self.assertContains(outcome_report, self.subject.name)

        other_level = AcademicLevel.objects.create(
            name="Grade 9 Remote",
            code="G9R",
            category="JUNIOR",
            order=90,
        )
        other_subject = LearningArea.objects.create(
            name="OTHER SCIENCE",
            code="OSCI",
        )
        other_subject.academic_levels.add(other_level)
        ELearningSubjectAllocation.objects.create(
            academic_level=other_level,
            learning_area=other_subject,
            teacher=self.other_teacher,
        )
        ELearningSubjectLessonPlan.objects.create(
            allocation=ELearningSubjectAllocation.objects.get(
                teacher=self.other_teacher,
                learning_area=other_subject,
            ),
            strand="Other strand only",
        )
        scoped_report = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "e-learning-reports"}),
            {
                "generate": "1",
                "report_date": "2026-08-21",
                "report_type": "lesson_plan",
                "level_id": str(self.level.id),
                "subject_id": str(self.subject.id),
            },
        )
        self.assertContains(scoped_report, self.subject.name)
        self.assertNotContains(scoped_report, "OTHER SCIENCE")
        self.assertNotContains(scoped_report, "Other strand only")
        self.assertNotContains(scoped_report, "Grade 9 Remote")

        assessments = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "assessments"})
        )
        self.assertEqual(assessments.status_code, 200)
        self.assertContains(assessments, "Register assessment")
        self.assertContains(assessments, "E-learning assessments")
        self.assertNotContains(assessments, "Tools coming soon")

        create = self.client.post(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "assessments"}),
            {
                "name": "E-learning midterm",
                "academic_year_id": str(self.year.id),
                "academic_term_id": str(self.term.id),
            },
        )
        from apps.curriculum.models import ELearningAssessment, ELearningAssessmentMark

        assessment = ELearningAssessment.objects.get(name="E-learning midterm")
        self.assertRedirects(
            create,
            reverse(
                "employees:teacher_elearning_assessment_detail",
                kwargs={"assessment_id": assessment.id},
            ),
        )
        detail = self.client.get(
            reverse(
                "employees:teacher_elearning_assessment_detail",
                kwargs={"assessment_id": assessment.id},
            )
        )
        self.assertEqual(detail.status_code, 200)
        self.assertContains(detail, "choose a class to enter marks")
        self.assertContains(detail, "Grade 1 East")
        self.assertContains(
            detail,
            reverse(
                "employees:teacher_elearning_assessment_class",
                kwargs={"assessment_id": assessment.id, "class_id": self.academic_class.id},
            ),
        )
        from apps.admissions.models import ParentGuardian, Student

        parent = ParentGuardian.objects.create(
            full_name="PAT ASSESS",
            relationship_to_student="MOTHER",
            phone_number="+254700008888",
        )
        student = Student.objects.create(
            first_name="BEN",
            last_name="LEARNER",
            date_of_birth="2018-01-01",
            gender=Student.Gender.MALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="EA-001",
            class_group="Grade 1 East",
            assessment_number="AEA001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        level_url = reverse(
            "employees:teacher_elearning_assessment_class",
            kwargs={"assessment_id": assessment.id, "class_id": self.academic_class.id},
        )
        level_sheet = self.client.get(level_url)
        self.assertEqual(level_sheet.status_code, 200)
        self.assertContains(level_sheet, "MATH")
        self.assertContains(level_sheet, "BEN LEARNER")
        self.assertContains(level_sheet, "Grade 1 East")
        self.assertContains(level_sheet, "Save marks")
        save_marks = self.client.post(
            level_url,
            {f"mark_{student.id}_{self.subject.id}": "40"},
        )
        self.assertRedirects(save_marks, level_url)
        mark = ELearningAssessmentMark.objects.get(
            assessment=assessment,
            student=student,
            learning_area=self.subject,
        )
        self.assertEqual(mark.marks, 40)

        level_page = self.client.get(
            reverse(
                "employees:teacher_elearning_attendance_level",
                kwargs={"level_id": self.level.id},
            )
        )
        self.assertEqual(level_page.status_code, 200)
        self.assertContains(level_page, "MATH")
        self.assertContains(level_page, "Attendance")
        profile_url = reverse(
            "employees:teacher_elearning_attendance_profile",
            kwargs={"level_id": self.level.id, "subject_id": self.subject.id},
        )
        self.assertContains(level_page, profile_url)

        from apps.admissions.models import ParentGuardian, Student
        from apps.curriculum.models import (
            ELearningAttendanceRecord,
            ELearningAttendanceSession,
        )

        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700009999",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="EL-001",
            class_group="Grade 1 East",
            assessment_number="AEL001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
            is_active=True,
        )
        profile = self.client.get(f"{profile_url}#attendance")
        self.assertEqual(profile.status_code, 200)
        self.assertContains(profile, "ANN EAST")
        self.assertContains(profile, "Save attendance")

        save = self.client.post(
            profile_url,
            {
                "form_action": "attendance",
                "lesson_date": "2026-08-21",
                "attendance_notes": "Online session",
                f"status_{student.id}": "ABSENT",
            },
        )
        self.assertRedirects(save, f"{profile_url}?date=2026-08-21#attendance")
        session = ELearningAttendanceSession.objects.get(
            allocation__academic_level=self.level,
            allocation__learning_area=self.subject,
            lesson_date=date(2026, 8, 21),
        )
        self.assertEqual(session.notes, "Online session")
        record = ELearningAttendanceRecord.objects.get(session=session, student=student)
        self.assertEqual(record.status, ELearningAttendanceRecord.Status.ABSENT)

    def test_teacher_can_upload_portal_ready_learning_materials(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        allocation = ELearningSubjectAllocation.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        self.client.login(username="135790", password="ReliablePass456")
        page = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "learning-materials"})
        )
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Upload learning material")
        self.assertContains(page, "Library")
        self.assertNotContains(page, "Tools coming soon")

        pdf = SimpleUploadedFile("week1-notes.pdf", b"%PDF-1.4 notes", content_type="application/pdf")
        cover = SimpleUploadedFile("cover.jpg", b"fake-image-bytes", content_type="image/jpeg")
        upload = self.client.post(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "learning-materials"}),
            {
                "form_action": "upload",
                "allocation": allocation.id,
                "content_format": ELearningLearningMaterial.ContentFormat.NOTES,
                "category": "Week 1",
                "name": "Fractions handout",
                "description": "Short notes for learners",
                "cover_image": cover,
                "material_file": pdf,
            },
        )
        self.assertRedirects(
            upload,
            reverse("employees:teacher_elearning_page", kwargs={"tool": "learning-materials"}),
        )
        material = ELearningLearningMaterial.objects.get(name="Fractions handout")
        self.assertEqual(material.category, "Week 1")
        self.assertEqual(material.file_extension, ".pdf")
        self.assertTrue(material.is_published)
        self.assertEqual(material.uploaded_by_id, self.teacher.id)

        listed = self.client.get(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "learning-materials"})
        )
        self.assertContains(listed, "Fractions handout")
        self.assertContains(listed, "Week 1")

        bad = SimpleUploadedFile("clip.mp3", b"ID3audio", content_type="audio/mpeg")
        rejected = self.client.post(
            reverse("employees:teacher_elearning_page", kwargs={"tool": "learning-materials"}),
            {
                "form_action": "upload",
                "allocation": allocation.id,
                "content_format": ELearningLearningMaterial.ContentFormat.NOTES,
                "category": "Week 2",
                "name": "Wrong format",
                "description": "",
                "material_file": bad,
            },
        )
        self.assertEqual(rejected.status_code, 200)
        self.assertFalse(ELearningLearningMaterial.objects.filter(name="Wrong format").exists())

    def test_exam_records_lists_all_exams(self):
        response = self.client.get(reverse("employees:teacher_exam_records"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Registered assessments")
        self.assertContains(response, "2026 TERM 3 assessment")
        self.assertContains(
            response,
            reverse("employees:teacher_exam_record_detail", kwargs={"exam_id": self.exam.id}),
        )
        self.assertNotContains(response, "Grade 1 East")
        self.assertNotContains(response, "No assessments are registered yet.")
        self.assertNotContains(response, "Session timetable")
        self.assertNotContains(response, reverse("employees:teacher_subject_attendance"))
        self.assertNotContains(response, "workspace-nav-label")

    def test_opening_an_exam_lists_allocated_subjects(self):
        response = self.client.get(
            reverse("employees:teacher_exam_record_detail", kwargs={"exam_id": self.exam.id})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Your subjects to teach")
        self.assertContains(response, "Subjects to teach")
        self.assertContains(response, "Mathematics")
        self.assertContains(response, "MATH")
        self.assertContains(response, "Assessment analytics")
        self.assertContains(
            response,
            reverse("employees:teacher_exam_analytics", kwargs={"exam_id": self.exam.id}),
        )
        self.assertContains(response, "Grade 1 East")
        self.assertContains(
            response,
            reverse(
                "employees:teacher_exam_record_class",
                kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
            ),
        )
        self.assertNotContains(response, "Grade 2 West")
        self.assertNotContains(response, "ENG2")
        self.assertNotContains(response, "Select a class from the sidebar")

    def test_teacher_exam_analytics_lists_allocated_classes_for_selection(self):
        other_stream = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )

        response = self.client.get(
            reverse("employees:teacher_exam_analytics", kwargs={"exam_id": self.exam.id})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Allocated classes")
        self.assertContains(response, "Other classes")
        self.assertContains(response, "All classes")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "Grade 1 West")
        self.assertContains(
            response,
            reverse("employees:teacher_exam_analytics_all", kwargs={"exam_id": self.exam.id}),
        )
        self.assertContains(
            response,
            reverse(
                "employees:teacher_exam_analytics_class",
                kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
            ),
        )
        self.assertContains(
            response,
            reverse(
                "employees:teacher_exam_analytics_class",
                kwargs={"exam_id": self.exam.id, "class_id": other_stream.id},
            ),
        )
        self.assertNotContains(response, "Grade 2 West")
        self.assertNotContains(response, "workspace-nav-label")

    def test_teacher_exam_analytics_all_classes_shows_every_allocated_class(self):
        from apps.admissions.models import ParentGuardian, Student

        second_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=second_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700001011",
            email="pat.all@example.com",
        )
        student_east = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1011",
            class_group="G1E",
            assessment_number="A1011",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        student_west = Student.objects.create(
            first_name="BEN",
            last_name="WEST",
            date_of_birth="2018-02-02",
            gender=Student.Gender.MALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1012",
            class_group="G1W",
            assessment_number="A1012",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=self.exam,
            student=student_east,
            learning_area=self.subject,
            marks=42,
        )
        ExamMark.objects.create(
            generation=self.exam,
            student=student_west,
            learning_area=self.subject,
            marks=31,
        )

        response = self.client.get(
            reverse("employees:teacher_exam_analytics_all", kwargs={"exam_id": self.exam.id})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "Grade 1 West")
        self.assertContains(response, "ANN EAST")
        self.assertContains(response, "BEN WEST")
        self.assertContains(response, "42")
        self.assertContains(response, "31")
        self.assertNotContains(response, "Grade 2 West")
        self.assertContains(
            response,
            reverse("employees:teacher_exam_analytics", kwargs={"exam_id": self.exam.id}),
        )
    def test_teacher_exam_analytics_shows_selected_class_results(self):
        from apps.admissions.models import ParentGuardian, Student

        second_class = AcademicClass.objects.create(
            academic_level=self.level,
            name="Grade 1 West",
            code="G1W",
            order=2,
        )
        ClassSubjectAllocation.objects.create(
            academic_class=second_class,
            learning_area=self.subject,
            teacher=self.teacher,
        )
        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700001010",
            email="pat.east@example.com",
        )
        student_east = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="G1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        student_west = Student.objects.create(
            first_name="BEN",
            last_name="WEST",
            date_of_birth="2018-02-02",
            gender=Student.Gender.MALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1002",
            class_group="G1W",
            assessment_number="A1002",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamMark.objects.create(
            generation=self.exam,
            student=student_east,
            learning_area=self.subject,
            marks=42,
        )
        ExamMark.objects.create(
            generation=self.exam,
            student=student_west,
            learning_area=self.subject,
            marks=31,
        )

        response = self.client.get(
            reverse(
                "employees:teacher_exam_analytics_class",
                kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Assessment analytics")
        self.assertContains(response, "Grade 1 East")
        self.assertContains(response, "ANN EAST")
        self.assertContains(response, "42")
        self.assertNotContains(response, "BEN WEST")
        self.assertNotContains(response, "31")
        self.assertContains(
            response,
            reverse(
                "employees:teacher_exam_record_class",
                kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
            ),
        )
        self.assertContains(
            response,
            reverse("employees:teacher_exam_analytics", kwargs={"exam_id": self.exam.id}),
        )

        denied = self.client.get(
            reverse(
                "employees:teacher_exam_analytics_class",
                kwargs={"exam_id": self.exam.id, "class_id": self.other_class.id},
            )
        )
        self.assertRedirects(
            denied,
            reverse("employees:teacher_exam_analytics", kwargs={"exam_id": self.exam.id}),
        )

    def test_teacher_can_view_selected_class_records(self):
        from apps.admissions.models import ParentGuardian, Student

        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700000555",
            email="pat.east@example.com",
        )
        Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        Student.objects.create(
            first_name="BEN",
            last_name="WEST",
            date_of_birth="2018-02-02",
            gender=Student.Gender.MALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1002",
            class_group="G1W",
            assessment_number="A1002",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        class_url = reverse(
            "employees:teacher_exam_record_class",
            kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
        )
        locked = self.client.get(class_url)
        self.assertEqual(locked.status_code, 200)
        self.assertContains(locked, "Grade 1 East")
        self.assertContains(locked, "ANN EAST")
        self.assertContains(locked, "1001")
        self.assertContains(locked, "Mathematics")
        self.assertContains(locked, "Locked · In session")
        self.assertContains(locked, "Marks are read-only")
        self.assertNotContains(locked, "data-exam-edit")
        self.assertNotContains(locked, "Save marks")
        self.assertContains(locked, "readonly")
        self.assertNotContains(locked, "BEN WEST")
        self.assertNotContains(locked, "English")

        self.exam.status = GeneratedExamTimetable.Status.MARKING
        self.exam.save(update_fields=["status"])
        editable = self.client.get(class_url)
        self.assertEqual(editable.status_code, 200)
        self.assertContains(editable, "data-exam-edit")
        self.assertContains(editable, "Save marks")
        self.assertNotContains(editable, "Marks are read-only")

    def test_teacher_cannot_save_marks_when_exam_is_not_marking(self):
        from apps.admissions.models import ParentGuardian, Student

        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700000555",
            email="pat.east@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        class_url = reverse(
            "employees:teacher_exam_record_class",
            kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
        )
        response = self.client.post(class_url, {f"mark_{student.id}_{self.subject.id}": "25"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Marks can only be edited while this assessment is in Marking status.")
        self.assertFalse(ExamMark.objects.filter(student=student, learning_area=self.subject).exists())

    def test_teacher_can_enter_marks_and_convert(self):
        from apps.admissions.models import ParentGuardian, Student

        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700000555",
            email="pat.east@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        ExamSubjectSetting.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            out_of_marks=50,
        )
        self.exam.status = GeneratedExamTimetable.Status.MARKING
        self.exam.save(update_fields=["status"])
        class_url = reverse(
            "employees:teacher_exam_record_class",
            kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
        )
        response = self.client.post(class_url, {f"mark_{student.id}_{self.subject.id}": "25"})
        self.assertRedirects(response, class_url)
        mark = ExamMark.objects.get(student=student, learning_area=self.subject)
        self.assertEqual(mark.marks, 25)
        self.assertEqual(mark.out_of_marks, 50)
        saved = self.client.get(class_url)
        self.assertContains(saved, 'value="25"')
        self.assertContains(saved, "50%")
        self.assertContains(saved, "Mean")
        self.assertContains(saved, "Student marks were saved.")

        invalid = self.client.post(class_url, {f"mark_{student.id}_{self.subject.id}": "80"})
        self.assertEqual(invalid.status_code, 200)
        self.assertContains(invalid, "within each subject's total marks")
        mark.refresh_from_db()
        self.assertEqual(mark.marks, 25)

    def test_changing_out_of_settings_does_not_alter_saved_percent_until_edit(self):
        from apps.admissions.models import ParentGuardian, Student

        parent = ParentGuardian.objects.create(
            full_name="PAT EAST",
            relationship_to_student="MOTHER",
            phone_number="+254700000555",
            email="pat.east@example.com",
        )
        student = Student.objects.create(
            first_name="ANN",
            last_name="EAST",
            date_of_birth="2018-01-01",
            gender=Student.Gender.FEMALE,
            academic_level=Student.AcademicLevel.GRADE_1,
            admission_number="1001",
            class_group="1E",
            assessment_number="A1001",
            sponsorship_category=Student.SponsorshipCategory.SELF,
            parent_guardian=parent,
        )
        setting = ExamSubjectSetting.objects.create(
            academic_level=self.level,
            learning_area=self.subject,
            out_of_marks=50,
        )
        self.exam.status = GeneratedExamTimetable.Status.MARKING
        self.exam.save(update_fields=["status"])
        class_url = reverse(
            "employees:teacher_exam_record_class",
            kwargs={"exam_id": self.exam.id, "class_id": self.academic_class.id},
        )
        self.client.post(class_url, {f"mark_{student.id}_{self.subject.id}": "25"})
        mark = ExamMark.objects.get(student=student, learning_area=self.subject)
        self.assertEqual(mark.marks, 25)
        self.assertEqual(mark.out_of_marks, 50)

        setting.out_of_marks = 100
        setting.save(update_fields=["out_of_marks", "updated_at"])

        page = self.client.get(class_url)
        self.assertContains(page, 'value="25"')
        self.assertContains(page, "50%")
        self.assertContains(page, 'data-out-of-changed="1"')
        self.assertContains(page, "Out-of marks settings have changed")
        self.assertContains(page, 'data-saved-out-of="50"')
        self.assertContains(page, 'data-current-out-of="100"')
        mark.refresh_from_db()
        self.assertEqual(mark.marks, 25)
        self.assertEqual(mark.out_of_marks, 50)

        save_response = self.client.post(class_url, {f"mark_{student.id}_{self.subject.id}": "25"})
        self.assertRedirects(save_response, class_url)
        mark.refresh_from_db()
        self.assertEqual(mark.marks, 25)
        self.assertEqual(mark.out_of_marks, 100)
        rebound = self.client.get(class_url)
        self.assertContains(rebound, "25%")
        self.assertNotContains(rebound, 'data-out-of-changed="1"')

    def test_teacher_cannot_open_a_class_that_is_not_allocated(self):
        response = self.client.get(
            reverse(
                "employees:teacher_exam_record_class",
                kwargs={"exam_id": self.exam.id, "class_id": self.other_class.id},
            )
        )
        self.assertRedirects(
            response,
            reverse("employees:teacher_exam_record_detail", kwargs={"exam_id": self.exam.id}),
        )

    def test_empty_exam_list_when_none_are_registered(self):
        self.exam.delete()
        response = self.client.get(reverse("employees:teacher_exam_records"))
        self.assertContains(response, "No assessments are registered yet.")

    def test_other_roles_cannot_open_teacher_exam_records(self):
        self.client.force_login(self.accountant)
        response = self.client.get(reverse("employees:teacher_exam_records"))
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "accountant"}),
        )
        response = self.client.get(
            reverse("employees:teacher_exam_record_detail", kwargs={"exam_id": self.exam.id})
        )
        self.assertRedirects(
            response,
            reverse("employees:role_dashboard", kwargs={"role": "accountant"}),
        )



